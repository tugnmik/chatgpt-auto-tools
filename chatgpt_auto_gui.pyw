"""
ChatGPT Auto Registration - MULTITHREADING VERSION (API-First)
Tự động đăng ký nhiều tài khoản ChatGPT đồng thời
Optimized: patchright + API calls thay vì Selenium UI automation
"""

import requests
import time
import re
import sys
import asyncio
import uuid
import random
import string
import json
import os
import threading
import concurrent.futures
from concurrent.futures import ThreadPoolExecutor, as_completed, TimeoutError as FutureTimeoutError
from colorama import init, Fore, Style
import warnings
import traceback
import tempfile
import shutil
import base64
import socket
import select
import pyotp
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from patchright.sync_api import sync_playwright
try:
    import tls_client
except ImportError:
    import subprocess as _sp
    _sp.check_call([sys.executable, "-m", "pip", "install", "tls-client"])
    import tls_client
import customtkinter as ctk
from tkinter import messagebox, filedialog
import tkinter.ttk as ttk

# Initialize colorama
init(autoreset=True)

# Locks for thread-safe operations
import atexit
import subprocess
file_lock = threading.Lock()
print_lock = threading.Lock()
driver_init_lock = threading.Lock()

# Global flag for getting checkout link (set by menu)
GET_CHECKOUT_LINK = False
GET_CHECKOUT_TYPE = "Plus"  # Options: "Plus", "Business", "Both"
ENABLE_2FA = False  # Global flag for enabling 2FA during registration

# Default password for registration (editable via GUI)
DEFAULT_PASSWORD = "Matkhau123!@#"

# Proxy settings
PROXY_CONFIG_FILE = "proxy_config.json"
PROXY_ENABLED = False
PROXY_STRING = ""
PROXY_FORMAT = "username:password@hostname:port"  # Default format

def detect_proxy_format(proxy_string):
    """Auto-detect proxy format from string.
    Supports:
      - username:password@hostname:port  (has @)
      - hostname:port:username:password   (port is 2nd, numeric)
      - username:password:hostname:port   (port is last, numeric)
    Returns format name or None."""
    proxy_string = proxy_string.strip()
    if not proxy_string:
        return None
    if '@' in proxy_string:
        return "username:password@hostname:port"
    parts = proxy_string.split(':')
    if len(parts) == 4:
        # Check if 2nd part is numeric → hostname:port:user:pass
        if parts[1].strip().isdigit():
            return "hostname:port:username:password"
        # Check if last part is numeric → user:pass:hostname:port
        if parts[3].strip().isdigit():
            return "username:password:hostname:port"
    return None

def parse_proxy(proxy_string, format_type=None):
    """Parse proxy string into components. Auto-detects format if not specified.
    Returns dict {host, port, username, password} and urls dict."""
    proxy_string = proxy_string.strip()
    if not proxy_string:
        return None, None
    
    if not format_type:
        format_type = detect_proxy_format(proxy_string)
    if not format_type:
        return None, None
    
    try:
        if format_type == "hostname:port:username:password":
            parts = proxy_string.split(':')
            if len(parts) != 4:
                return None, None
            host, port, username, password = parts
        elif format_type == "username:password:hostname:port":
            parts = proxy_string.split(':')
            if len(parts) != 4:
                return None, None
            username, password, host, port = parts
        elif format_type == "username:password@hostname:port":
            if '@' not in proxy_string:
                return None, None
            creds, server = proxy_string.split('@', 1)
            cred_parts = creds.split(':', 1)
            server_parts = server.split(':', 1)
            if len(cred_parts) != 2 or len(server_parts) != 2:
                return None, None
            username, password = cred_parts
            host, port = server_parts
        else:
            return None, None
        
        port = port.strip()
        proxy_info = {
            "host": host.strip(),
            "port": port,
            "username": username.strip(),
            "password": password.strip()
        }
        chrome_url = f"http://{host.strip()}:{port}"
        requests_url = f"http://{username.strip()}:{password.strip()}@{host.strip()}:{port}"
        return proxy_info, {"chrome": chrome_url, "requests": requests_url}
    except Exception:
        return None, None

class LocalProxyAuthBridge:
    """Local HTTP proxy bridge (no auth) -> upstream HTTP proxy (with Basic auth).

    Why: Chrome + undetected_chromedriver is unreliable with proxy auth popups / extensions / CDP.
    This bridge makes Chrome talk to 127.0.0.1 without authentication, while the bridge
    injects Proxy-Authorization for the upstream proxy.
    """

    def __init__(self, upstream_host, upstream_port, username, password, log_func=None):
        self.upstream_host = (upstream_host or "").strip()
        self.upstream_port = int(upstream_port)
        self.username = username or ""
        self.password = password or ""
        self._log = log_func
        self._stop_event = threading.Event()
        self._server = None
        self._thread = None
        self.bound_host = "127.0.0.1"
        self.bound_port = None

    def start(self):
        if self._thread and self._thread.is_alive():
            return

        server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        server.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        server.bind((self.bound_host, 0))
        server.listen(50)
        self._server = server
        self.bound_port = int(server.getsockname()[1])

        self._thread = threading.Thread(target=self._serve, name="LocalProxyAuthBridge", daemon=True)
        self._thread.start()

        self._safe_log(f"[ProxyBridge] {self.bound_host}:{self.bound_port} -> {self.upstream_host}:{self.upstream_port}")

    def stop(self):
        self._stop_event.set()
        try:
            if self._server:
                self._server.close()
        except Exception:
            pass
        self._server = None

    def _safe_log(self, msg):
        if not self._log:
            return
        try:
            self._log(msg)
        except Exception:
            pass

    def _serve(self):
        while not self._stop_event.is_set() and self._server:
            try:
                client, _addr = self._server.accept()
                threading.Thread(target=self._handle_client, args=(client,), daemon=True).start()
            except OSError:
                break
            except Exception:
                continue

    def _recv_until(self, sock_obj, marker=b"\r\n\r\n", max_bytes=256 * 1024):
        data = b""
        while marker not in data and len(data) < max_bytes:
            chunk = sock_obj.recv(4096)
            if not chunk:
                break
            data += chunk
        return data

    def _inject_proxy_auth_header(self, header_bytes):
        auth_b64 = base64.b64encode(f"{self.username}:{self.password}".encode("utf-8")).decode("ascii")
        auth_line = f"Proxy-Authorization: Basic {auth_b64}".encode("latin1")

        low = header_bytes.lower()
        if b"\r\nproxy-authorization:" in low:
            return header_bytes

        # Insert after request line
        parts = header_bytes.split(b"\r\n", 1)
        if len(parts) != 2:
            return header_bytes
        return parts[0] + b"\r\n" + auth_line + b"\r\n" + parts[1]

    def _tunnel(self, sock_a, sock_b):
        sock_a.setblocking(False)
        sock_b.setblocking(False)
        sockets = [sock_a, sock_b]
        while True:
            try:
                readable, _w, _x = select.select(sockets, [], [], 30)
            except Exception:
                return
            if not readable:
                continue
            for s in readable:
                other = sock_b if s is sock_a else sock_a
                try:
                    data = s.recv(65536)
                except Exception:
                    return
                if not data:
                    return
                try:
                    other.sendall(data)
                except Exception:
                    return

    def _handle_client(self, client):
        upstream = None
        try:
            client.settimeout(15)
            header = self._recv_until(client)
            if not header:
                return

            upstream = socket.create_connection((self.upstream_host, self.upstream_port), timeout=20)
            upstream.settimeout(20)

            header = self._inject_proxy_auth_header(header)
            upstream.sendall(header)

            # For CONNECT, we need to forward the upstream response header first, then tunnel.
            first_line = header.split(b"\r\n", 1)[0].decode("latin1", errors="ignore")
            if first_line.upper().startswith("CONNECT "):
                resp = self._recv_until(upstream, max_bytes=64 * 1024)
                if not resp:
                    return
                client.sendall(resp)
                # Only tunnel on 200
                status_line = resp.split(b"\r\n", 1)[0]
                if b" 200 " not in status_line:
                    return

            # Tunnel both directions (also works for non-CONNECT)
            self._tunnel(client, upstream)
        except Exception:
            return
        finally:
            try:
                client.close()
            except Exception:
                pass
            try:
                if upstream:
                    upstream.close()
            except Exception:
                pass

def get_proxy_for_requests():
    """Get proxy dict for requests library. Returns None if proxy disabled."""
    if not PROXY_ENABLED or not PROXY_STRING:
        return None
    proxy_info, urls = parse_proxy(PROXY_STRING)
    if not proxy_info or not urls:
        return None
    req_url = urls["requests"]
    return {"http": req_url, "https": req_url}

def load_proxy_config():
    """Load proxy configuration from JSON file."""
    global PROXY_ENABLED, PROXY_STRING, PROXY_FORMAT
    try:
        if os.path.exists(PROXY_CONFIG_FILE):
            with open(PROXY_CONFIG_FILE, 'r', encoding='utf-8') as f:
                config = json.load(f)
            PROXY_ENABLED = False  # Always start with proxy disabled
            PROXY_STRING = config.get("proxy_string", "")
            PROXY_FORMAT = config.get("format", "username:password@hostname:port")
            return PROXY_ENABLED, PROXY_STRING, PROXY_FORMAT
    except Exception:
        pass
    return False, "", "username:password@hostname:port"

def save_proxy_config(enabled, proxy_string, fmt):
    """Save proxy configuration to JSON file."""
    try:
        config = {
            "enabled": enabled,
            "proxy_string": proxy_string,
            "format": fmt
        }
        with open(PROXY_CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
        return True
    except Exception:
        return False

# Load proxy config on startup
load_proxy_config()

# Define colors
class Colors:
    SUCCESS = Fore.GREEN
    ERROR = Fore.RED
    WARNING = Fore.YELLOW
    INFO = Fore.CYAN
    HEADER = Fore.MAGENTA
    RESET = Style.RESET_ALL


def safe_print(thread_id, message, color=Colors.INFO, emoji=""):
    """Thread-safe print with thread ID"""
    with print_lock:
        print(f"[T{thread_id}] {emoji}{color}{message}{Colors.RESET}")


class TempMailAPI:
    """API client cho tinyhost.shop"""
    
    def __init__(self):
        self.base_url = "https://tinyhost.shop"
    
    @property
    def proxies(self):
        """Always use latest proxy settings"""
        return get_proxy_for_requests()
    
    def get_random_domains(self, limit=10):
        """Lấy danh sách domain ngẫu nhiên"""
        try:
            url = f"{self.base_url}/api/random-domains/"
            params = {"limit": limit}
            response = requests.get(url, params=params, timeout=10, proxies=self.proxies)
            response.raise_for_status()
            return response.json()
        except Exception as e:
            print(f"[TinyHost] get_random_domains error: {e}")
            return None
    
    def get_emails(self, domain, user, page=1, limit=20):
        """Lấy danh sách email"""
        try:
            url = f"{self.base_url}/api/email/{domain}/{user}/"
            params = {"page": page, "limit": limit}
            response = requests.get(url, params=params, timeout=10, proxies=self.proxies)
            response.raise_for_status()
            return response.json()
        except Exception as e:
            print(f"[TinyHost] get_emails error: {e}")
            return None
    
    def get_email_detail(self, domain, user, email_id):
        """Lấy chi tiết một email"""
        try:
            url = f"{self.base_url}/api/email/{domain}/{user}/{email_id}"
            response = requests.get(url, timeout=10, proxies=self.proxies)
            response.raise_for_status()
            return response.json()
        except Exception as e:
            print(f"[TinyHost] get_email_detail error: {e}")
            return None
    
    def generate_email(self):
        """Tạo một email ngẫu nhiên"""
        domains_data = self.get_random_domains(20)
        if not domains_data or not domains_data.get('domains'):
            print(f"[TinyHost] generate_email: no domains! Response: {domains_data}")
            return None
        
        domain = random.choice(domains_data['domains'])
        username = ''.join(random.choices(string.ascii_lowercase + string.digits, k=10))
        email = f"{username}@{domain}"
        
        return {
            'email': email,
            'username': username,
            'domain': domain
        }


class DongVanOAuth2API:
    """API client cho DongVan OAuth2 (thay thế tinyhost.shop)"""
    
    API_MESSAGES_URL = "https://tools.dongvanfb.net/api/get_messages_oauth2"
    
    def __init__(self, email, password, refresh_token, client_id):
        self.email = email
        self.password = password
        self.refresh_token = refresh_token
        self.client_id = client_id
    
    def fetch_messages(self):
        """Lấy danh sách email từ OAuth2 API"""
        payload = {
            "email": self.email,
            "refresh_token": self.refresh_token,
            "client_id": self.client_id,
        }
        try:
            proxies = get_proxy_for_requests()
            response = requests.post(self.API_MESSAGES_URL, json=payload, timeout=20, proxies=proxies)
            response.raise_for_status()
            return response.json()
        except requests.RequestException:
            return None
        except json.JSONDecodeError:
            return None
    
    def extract_code_from_messages(self, messages_payload):
        """Trích xuất code 6 số từ thư mới nhất"""
        if not messages_payload:
            return None
        
        messages = messages_payload.get("messages")
        if not isinstance(messages, list) or not messages:
            return None
        
        pattern = re.compile(r"\b(\d{6})\b")
        
        def parse_msg_datetime(raw):
            if not raw:
                return datetime.min
            try:
                return datetime.fromisoformat(raw.replace("Z", "+00:00"))
            except ValueError:
                pass
            for fmt in ("%H:%M - %d/%m/%Y", "%d/%m/%Y %H:%M:%S"):
                try:
                    return datetime.strptime(raw, fmt)
                except ValueError:
                    continue
            return datetime.min
        
        # Sắp xếp thư mới nhất lên đầu
        sorted_messages = sorted(
            messages,
            key=lambda msg: parse_msg_datetime(msg.get("date")),
            reverse=True,
        )
        
        if not sorted_messages:
            return None
        
        latest_msg = sorted_messages[0]
        
        # Ưu tiên 1: Trường 'code' nếu có
        code_field = latest_msg.get("code", "")
        if code_field and pattern.match(str(code_field)):
            return code_field
        
        # Ưu tiên 2: Extract từ subject line
        subject = latest_msg.get("subject") or ""
        subject_codes = pattern.findall(subject)
        if subject_codes:
            return subject_codes[0]
        
        # Ưu tiên 3: Extract từ content/message
        content = latest_msg.get("content") or latest_msg.get("message") or ""
        content_codes = pattern.findall(content)
        if content_codes:
            return content_codes[0]
        
        return None
    
    def get_email_info(self):
        """Trả về thông tin email theo format tương thích"""
        return {
            'email': self.email,
            'username': self.email.split('@')[0],
            'domain': self.email.split('@')[1] if '@' in self.email else ''
        }


# Global list để quản lý tài khoản OAuth2
oauth2_accounts = []
account_lock = threading.Lock()


def load_oauth2_accounts_from_excel(file_path="oauth2.xlsx", skip_registered=True):
    """Load tài khoản OAuth2 từ file Excel oauth2.xlsx
    Cột A: email|password|refresh_token|client_id
    Cột B: Status (registered = đã đăng ký, trống = chưa đăng ký)
    Row 1 là header, dữ liệu từ row 2
    
    Args:
        file_path: Path to oauth2.xlsx
        skip_registered: If True, skip accounts with status='registered' (for Registration)
                        If False, load all accounts (for MFA/Checkout lookup)
    """
    accounts = []
    
    if not os.path.exists(file_path):
        return accounts
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        skipped_count = 0
        
        for row_num in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_num, column=1).value
            if not cell_value:
                continue
            
            line = str(cell_value).strip()
            if not line or line.startswith("#"):
                continue
            
            parts = [part.strip() for part in line.split("|")]
            if len(parts) != 4:
                continue
            
            # Kiểm tra cột B (Status) - skip nếu đã registered (chỉ khi skip_registered=True)
            if skip_registered:
                status_value = ws.cell(row=row_num, column=2).value
                if status_value and str(status_value).strip().lower() == "registered":
                    skipped_count += 1
                    continue
            
            email, password, refresh_token, client_id = parts
            accounts.append({
                "email": email,
                "password": password,
                "refresh_token": refresh_token,
                "client_id": client_id,
                "row_num": row_num,
                "used": False
            })
        
        wb.close()
        
        if skipped_count > 0 and skip_registered:
            print(f"⏭️ Skipped {skipped_count} registered accounts")
        
    except Exception as e:
        print(f"Error loading OAuth2 accounts: {e}")
        return []
    
    return accounts


def get_next_oauth2_account():
    """Lấy tài khoản OAuth2 tiếp theo chưa được sử dụng (thread-safe)"""
    global oauth2_accounts
    
    with account_lock:
        for i, account in enumerate(oauth2_accounts):
            if not account.get("used", False):
                oauth2_accounts[i]["used"] = True
                return account
    return None


def reset_oauth2_accounts():
    """Reset trạng thái used của tất cả oauth2 accounts"""
    global oauth2_accounts
    with account_lock:
        for i in range(len(oauth2_accounts)):
            oauth2_accounts[i]["used"] = False


def mark_oauth2_registered(row_num, file_path="oauth2.xlsx"):
    """Ghi 'registered' vào cột B của oauth2.xlsx sau khi đăng ký thành công
    Thread-safe using file_lock
    """
    with file_lock:
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            ws.cell(row=row_num, column=2, value="registered")
            wb.save(file_path)
            wb.close()
            return True
        except Exception as e:
            print(f"Error marking OAuth2 account as registered: {e}")
            return False


class ChatGPTAutoRegisterWorker:
    """Worker thread for ChatGPT registration (API-First with patchright)"""
    
    def __init__(self, thread_id, num_threads=1, email_mode="TinyHost", oauth2_account=None):
        self.thread_id = thread_id
        self.num_threads = num_threads
        self.email_mode = email_mode
        self.oauth2_account = oauth2_account
        
        # Initialize mail API based on mode
        if email_mode == "OAuth2" and oauth2_account:
            self.mail_api = DongVanOAuth2API(
                email=oauth2_account["email"],
                password=oauth2_account["password"],
                refresh_token=oauth2_account["refresh_token"],
                client_id=oauth2_account["client_id"]
            )
            self.email_info = self.mail_api.get_email_info()
            self.password = DEFAULT_PASSWORD
            self.oauth2_row_num = oauth2_account.get("row_num")
        else:
            self.mail_api = TempMailAPI()
            self.email_info = None
            self.password = DEFAULT_PASSWORD
            self.oauth2_row_num = None
        
        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None
        self.proxy_bridge = None
        self.stop_event = None
        self._pw_server_pid = None  # Playwright server PID for force kill
        self.max_retries = 2
        self.current_retry = 0
        
    def log(self, message, color=Colors.INFO, emoji=""):
        """Log with thread ID"""
        safe_print(self.thread_id, message, color, emoji)

    def cleanup_browser(self):
        """Close patchright browser and cleanup"""
        # If stop was requested, kill by PID (fast) instead of Playwright close (hangs)
        is_stopping = self.stop_event and self.stop_event.is_set()
        
        if is_stopping:
            # Kill Playwright server PID + all children (browsers)
            if self._pw_server_pid:
                try:
                    subprocess.Popen(
                        ['taskkill', '/F', '/PID', str(self._pw_server_pid), '/T'],
                        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                        creationflags=subprocess.CREATE_NO_WINDOW
                    )
                except Exception:
                    pass
            # Just nullify everything
            self.page = None
            self.context = None
            self.browser = None
            self.playwright = None
            self._pw_server_pid = None
            if self.proxy_bridge:
                try:
                    self.proxy_bridge.stop()
                except Exception:
                    pass
                self.proxy_bridge = None
            return
        
        # Normal cleanup (not stopping) — use Playwright close methods
        if self.page:
            try:
                self.page.close()
            except Exception:
                pass
            self.page = None
        if self.context:
            try:
                self.context.close()
            except Exception:
                pass
            self.context = None
        if self.browser:
            try:
                browser_pid = None
                try:
                    browser_pid = self.browser.process.pid if hasattr(self.browser, 'process') and self.browser.process else None
                except Exception:
                    pass
                self.browser.close()
                if browser_pid:
                    try:
                        import signal
                        os.kill(browser_pid, signal.SIGTERM)
                    except (ProcessLookupError, OSError):
                        pass
            except Exception:
                pass
            self.browser = None
        if self.playwright:
            try:
                self.playwright.stop()
            except Exception:
                pass
            self.playwright = None
        if self.proxy_bridge:
            try:
                self.proxy_bridge.stop()
            except Exception:
                pass
            self.proxy_bridge = None

    def setup_browser(self, max_retries=3):
        """Initialize patchright browser with retry"""
        for attempt in range(max_retries):
            try:
                if self.stop_event and self.stop_event.is_set():
                    return False
                
                self.log(f"Initializing browser (attempt {attempt + 1}/{max_retries})...", Colors.INFO, "🔄 ")
                
                self.playwright = sync_playwright().start()
                
                # Capture Playwright server PID for reliable force kill
                try:
                    self._pw_server_pid = self.playwright._impl_obj._connection._transport._proc.pid
                except Exception:
                    self._pw_server_pid = None
                
                # Build launch args
                launch_args = [
                    '--no-sandbox',
                    '--disable-dev-shm-usage',
                    '--disable-blink-features=AutomationControlled',
                    '--lang=en-US',
                    '--disable-background-timer-throttling',
                    '--disable-backgrounding-occluded-windows',
                    '--disable-renderer-backgrounding',
                ]
                
                # Apply proxy if enabled
                proxy_config = None
                if PROXY_ENABLED and PROXY_STRING:
                    proxy_info, urls = parse_proxy(PROXY_STRING)
                    if proxy_info and urls:
                        host = (proxy_info.get("host") or "").strip()
                        port = str(proxy_info.get("port") or "").strip()
                        username = (proxy_info.get("username") or "").strip()
                        password_p = (proxy_info.get("password") or "").strip()
                        if host and port.isdigit():
                            if username and password_p:
                                proxy_config = {
                                    "server": f"http://{host}:{port}",
                                    "username": username,
                                    "password": password_p,
                                }
                            else:
                                proxy_config = {"server": f"http://{host}:{port}"}
                            self.log(f"Proxy enabled: {host}:{port}", Colors.INFO, "🌐 ")
                
                self.browser = self.playwright.chromium.launch(
                    headless=False,
                    args=launch_args,
                    proxy=proxy_config,
                )
                
                # Window positioning for multi-thread
                viewport_w, viewport_h = 1000, 750
                position_index = (self.thread_id - 1) % self.num_threads if self.num_threads > 1 else 0
                
                self.context = self.browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/145.0.0.0 Safari/537.36",
                    viewport={"width": viewport_w, "height": viewport_h},
                )
                self.page = self.context.new_page()
                
                self.log("Browser initialized!", Colors.SUCCESS, "✅ ")
                return True
                
            except Exception as e:
                self.log(f"Error initializing browser (attempt {attempt + 1}): {e}", Colors.WARNING, "⚠️ ")
                self.cleanup_browser()
                if attempt < max_retries - 1:
                    time.sleep(attempt * 2 + 1)
                    continue
                else:
                    self.log("Could not initialize browser", Colors.ERROR, "❌ ")
                    return False
        return False

    @staticmethod
    def _random_name():
        first = ["James","Robert","John","Michael","David","William","Richard",
                 "Joseph","Thomas","Christopher","Daniel","Matthew","Anthony",
                 "Mary","Patricia","Jennifer","Linda","Elizabeth","Jessica","Sarah"]
        last = ["Smith","Johnson","Williams","Brown","Jones","Garcia","Miller",
                "Davis","Rodriguez","Martinez","Wilson","Anderson","Thomas","Taylor"]
        return random.choice(first), random.choice(last)

    @staticmethod
    def _random_birthdate():
        start = datetime(1990, 1, 1)
        end = datetime(2005, 12, 31)
        days = random.randint(0, (end - start).days)
        return (start + timedelta(days=days)).strftime("%Y-%m-%d")

    def _wait_for_otp_tinyhost(self, email, timeout=120):
        """Wait for OTP from TinyHost email"""
        mail_api = self.mail_api
        start_time = time.time()
        checked_ids = set()
        pattern = re.compile(r"\b(\d{6})\b")
        
        while time.time() - start_time < timeout:
            if self.stop_event and self.stop_event.is_set():
                return None
            try:
                domain = email.split("@")[1]
                user = email.split("@")[0]
                data = mail_api.get_emails(domain, user, limit=10)
                if data:
                    emails_list = data.get("emails", [])
                    for mail in emails_list:
                        mail_id = mail.get("id")
                        if mail_id in checked_ids:
                            continue
                        
                        sender = mail.get("sender", "")
                        subject = mail.get("subject", "")
                        
                        # Filter for OpenAI emails
                        if "openai" not in sender.lower() and "openai" not in subject.lower():
                            checked_ids.add(mail_id)
                            continue
                        
                        self.log(f"Checking email [{mail_id}]: '{subject}' from {sender}", Colors.INFO)
                        
                        # Try extract from subject
                        codes = pattern.findall(subject)
                        if codes:
                            self.log(f"OTP from subject: {codes[0]}", Colors.SUCCESS, "✅ ")
                            return codes[0]
                        
                        # Try extract from body
                        body = mail.get("body", "") or ""
                        html_body = mail.get("html_body", "") or ""
                        for text in [body, html_body]:
                            codes = pattern.findall(text)
                            if codes:
                                self.log(f"OTP from body: {codes[0]}", Colors.SUCCESS, "✅ ")
                                return codes[0]
                        
                        # Try full detail
                        try:
                            detail = mail_api.get_email_detail(domain, user, mail_id)
                            if detail:
                                for field in ["subject", "body", "html_body"]:
                                    text = detail.get(field, "") or ""
                                    codes = pattern.findall(text)
                                    if codes:
                                        self.log(f"OTP from detail.{field}: {codes[0]}", Colors.SUCCESS, "✅ ")
                                        return codes[0]
                        except Exception:
                            pass
                        
                        checked_ids.add(mail_id)
            except Exception as e:
                self.log(f"Email poll error: {e}", Colors.WARNING)
            
            remaining = int(timeout - (time.time() - start_time))
            if remaining > 0 and remaining % 10 == 0:
                self.log(f"Waiting for OTP... ({remaining}s remaining)", Colors.INFO, "⏳ ")
            time.sleep(1)
        
        self.log(f"OTP timeout after {timeout}s", Colors.ERROR, "❌ ")
        return None

    def _wait_for_otp_oauth2(self, timeout=120):
        """Wait for OTP from OAuth2 email"""
        start_time = time.time()
        last_code = None
        
        while time.time() - start_time < timeout:
            if self.stop_event and self.stop_event.is_set():
                return None
            try:
                messages = self.mail_api.fetch_messages()
                if messages:
                    code = self.mail_api.extract_code_from_messages(messages)
                    if code and code != last_code:
                        self.log(f"OTP from OAuth2: {code}", Colors.SUCCESS, "✅ ")
                        return code
                    last_code = code
            except Exception as e:
                self.log(f"OAuth2 poll error: {e}", Colors.WARNING)
            
            remaining = int(timeout - (time.time() - start_time))
            if remaining > 0 and remaining % 10 == 0:
                self.log(f"Waiting for OTP... ({remaining}s remaining)", Colors.INFO, "⏳ ")
            time.sleep(2)
        
        self.log(f"OTP timeout after {timeout}s", Colors.ERROR, "❌ ")
        return None

    def _setup_2fa_api(self, access_token):
        """API-based 2FA setup using page.evaluate(fetch())"""
        page = self.page
        self.log("[2FA] Starting 2FA setup via API...", Colors.INFO, "🔐 ")
        
        # Step 1: Check current MFA status
        mfa_info = page.evaluate("""
            async (token) => {
                const r = await fetch('/backend-api/accounts/mfa_info', {
                    headers: { 'Authorization': 'Bearer ' + token }
                });
                return { status: r.status, body: await r.text() };
            }
        """, access_token)
        
        try:
            info_data = json.loads(mfa_info.get("body", "{}"))
            if info_data.get("totp", {}).get("is_enabled"):
                self.log("[2FA] Already enabled!", Colors.WARNING, "⚠️ ")
                return None
        except Exception:
            pass
        
        # Step 2: Enroll TOTP
        self.log("[2FA] Enrolling TOTP...", Colors.INFO)
        enroll_resp = page.evaluate("""
            async (token) => {
                const r = await fetch('/backend-api/accounts/mfa/enroll', {
                    method: 'POST',
                    headers: {
                        'Authorization': 'Bearer ' + token,
                        'Content-Type': 'application/json'
                    },
                    body: '{"factor_type":"totp"}'
                });
                return { status: r.status, body: await r.text() };
            }
        """, access_token)
        
        if enroll_resp.get("status") != 200:
            self.log(f"[2FA] Enroll failed ({enroll_resp.get('status')})", Colors.ERROR, "❌ ")
            return None
        
        try:
            enroll_data = json.loads(enroll_resp.get("body", "{}"))
        except Exception:
            self.log("[2FA] Failed to parse enroll response", Colors.ERROR, "❌ ")
            return None
        
        # Extract secret
        secret = enroll_data.get("secret") or enroll_data.get("totp_secret") or ""
        totp_url = enroll_data.get("barcode_uri") or enroll_data.get("totp_url") or ""
        recovery_codes = enroll_data.get("recovery_codes") or []
        
        if not secret and totp_url and "secret=" in totp_url:
            secret = totp_url.split("secret=")[1].split("&")[0]
        
        if not secret:
            self.log(f"[2FA] No secret found. Keys: {list(enroll_data.keys())}", Colors.ERROR, "❌ ")
            return None
        
        self.log(f"[2FA] Secret: {secret}", Colors.INFO)
        
        # Step 3: Generate TOTP code
        totp = pyotp.TOTP(secret)
        code = totp.now()
        self.log(f"[2FA] TOTP code: {code}", Colors.INFO)
        
        # Step 4: Activate enrollment
        session_id = enroll_data.get("session_id", "")
        factor_type = enroll_data.get("factor", {}).get("factor_type", "totp")
        activate_payload = json.dumps({
            "code": code,
            "session_id": session_id,
            "factor_type": factor_type,
        })
        
        activate_resp = page.evaluate("""
            async ([token, payload]) => {
                const r = await fetch('/backend-api/accounts/mfa/user/activate_enrollment', {
                    method: 'POST',
                    headers: {
                        'Authorization': 'Bearer ' + token,
                        'Content-Type': 'application/json'
                    },
                    body: payload
                });
                return { status: r.status, body: await r.text() };
            }
        """, [access_token, activate_payload])
        
        if activate_resp.get("status") != 200:
            self.log(f"[2FA] Activation failed ({activate_resp.get('status')})", Colors.ERROR, "❌ ")
            return None
        
        self.log(f"[2FA] ✅ 2FA enabled! Secret: {secret}", Colors.SUCCESS, "🔐 ")
        return {
            "secret": secret,
            "totp_url": totp_url,
            "recovery_codes": recovery_codes,
        }

    def _call_checkout_api(self, access_token, payload):
        """Gọi POST /backend-api/payments/checkout → trả về checkout URL"""
        return call_checkout_api(access_token, payload, "Checkout", self.log)
    
    # Checkout payloads by plan type
    CHECKOUT_PAYLOADS = {
        "Plus": {
            "plan_name": "chatgptplusplan",
            "billing_details": {"country": "VN", "currency": "VND"},
            "checkout_ui_mode": "custom",
            "promo_campaign": {"promo_campaign_id": "plus-1-month-free", "is_coupon_from_query_param": False}
        },
        "Business": {
            "plan_name": "chatgptteamplan",
            "team_plan_data": {
                "workspace_name": "SABUBULEX",
                "price_interval": "month",
                "seat_quantity": 5
            },
            "billing_details": {"country": "VN", "currency": "VND"},
            "checkout_ui_mode": "custom",
            "promo_campaign": {
                "promo_campaign_id": "team-1-month-free",
                "is_coupon_from_query_param": True
            }
        },
    }

    def get_checkout_link_via_api(self, access_token, plan_type="Plus"):
        """Lấy checkout link qua API. plan_type: 'Plus' hoặc 'Business'"""
        emoji = "🔗 " if plan_type == "Plus" else "💼 "
        try:
            self.log(f"Lấy {plan_type} checkout link...", Colors.INFO, emoji)
            if not access_token:
                return None

            payload = self.CHECKOUT_PAYLOADS.get(plan_type)
            if not payload:
                self.log(f"Unknown plan type: {plan_type}", Colors.ERROR, "❌ ")
                return None

            url = self._call_checkout_api(access_token, payload)

            if url:
                self.log(f"{plan_type} Checkout URL (API): {url}", Colors.SUCCESS, "✅ ")
            else:
                self.log(f"Không lấy được {plan_type} link qua API", Colors.WARNING, "⚠️ ")

            return url
        except Exception as e:
            self.log(f"API {plan_type} checkout error: {e}", Colors.ERROR, "❌ ")
            return None

    def save_account_info(self, session_data, checkout_url=None, business_checkout_url=None, mfa_secret=None):
        """Save account info to Excel (thread-safe)"""
        try:
            if not self.email_info:
                return False
            
            if not session_data:
                self.log("No session data to save", Colors.WARNING, "⚠️ ")
                return False

            filename = "chatgpt.xlsx"
            account = f"{self.email_info['email']}:{self.password}"
            
            with file_lock:
                if os.path.exists(filename):
                    wb = load_workbook(filename)
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Accounts"
                    ws['A1'] = "Account"
                    ws['B1'] = "Session JSON"
                    ws['C1'] = "Plus Checkout URL"
                    ws['D1'] = "Business Checkout URL"
                    ws['E1'] = "2FA Secret"
                
                if ws['D1'].value is None:
                    ws['D1'] = "Business Checkout URL"
                if ws['E1'].value is None:
                    ws['E1'] = "2FA Secret"
                
                next_row = ws.max_row + 1
                ws[f'A{next_row}'] = account
                ws[f'B{next_row}'] = session_data
                ws[f'C{next_row}'] = checkout_url if checkout_url else ""
                ws[f'D{next_row}'] = business_checkout_url if business_checkout_url else ""
                ws[f'E{next_row}'] = mfa_secret if mfa_secret else ""
                
                wb.save(filename)
                wb.close()
            
            self.log("Info saved to chatgpt.xlsx", Colors.SUCCESS, "✅ ")
            if checkout_url:
                self.log("Plus Checkout URL saved!", Colors.SUCCESS, "✅ ")
            if business_checkout_url:
                self.log("Business Checkout URL saved!", Colors.SUCCESS, "💼 ")
            if mfa_secret:
                self.log("2FA Secret saved!", Colors.SUCCESS, "🔐 ")
            return True
            
        except Exception as e:
            self.log(f"Error saving info: {e}", Colors.ERROR, "❌ ")
            return False

    def run(self):
        """Run the entire registration flow (API-First)"""
        while self.current_retry <= self.max_retries:
            try:
                # Check stop at start of each retry
                if self.stop_event and self.stop_event.is_set():
                    self.cleanup_browser()
                    return (False, None)
                
                if self.current_retry > 0:
                    self.log(f"Retry attempt {self.current_retry}/{self.max_retries}...", Colors.WARNING, "🔄 ")
                    time.sleep(3)
                
                self.log("─" * 40, Colors.HEADER)
                self.log("STARTING REGISTRATION (API-First)", Colors.HEADER, "🚀 ")
                self.log("─" * 40, Colors.HEADER)
                
                # === Step 0: Generate email if TinyHost ===
                if self.email_mode != "OAuth2":
                    email_data = self.mail_api.generate_email()
                    if not email_data:
                        self.log("Failed to generate email", Colors.ERROR, "❌ ")
                        return (False, None)
                    self.email_info = email_data
                
                email = self.email_info['email']
                password = self.password
                
                first_name, last_name = self._random_name()
                full_name = f"{first_name} {last_name}"
                birthdate = self._random_birthdate()
                device_id = str(uuid.uuid4())
                
                self.log(f"Email: {email} | Pass: {password}", Colors.INFO, "📧 ")
                self.log(f"Name: {full_name} | DOB: {birthdate}", Colors.INFO, "👤 ")
                
                # === Step 1: Setup browser ===
                if not self.setup_browser():
                    return (False, None)
                
                page = self.page
                
                # === Step 2: Navigate + CSRF ===
                self.log("Navigating to chatgpt.com...", Colors.INFO, "🌐 ")
                page.goto("https://chatgpt.com/", wait_until="networkidle", timeout=60000)
                if self.stop_event and self.stop_event.is_set():
                    self.cleanup_browser()
                    return (False, None)
                page.wait_for_timeout(2000)
                
                # === Step 3: Get CSRF token ===
                self.log("Getting CSRF token...", Colors.INFO, "🔑 ")
                csrf_resp = page.evaluate("""
                    async () => {
                        const r = await fetch('/api/auth/csrf', { headers: { 'Content-Type': 'application/json' } });
                        return await r.json();
                    }
                """)
                csrf_token = csrf_resp.get("csrfToken", "")
                if not csrf_token:
                    self.log("Failed to get CSRF token", Colors.ERROR, "❌ ")
                    self.cleanup_browser()
                    self.current_retry += 1
                    continue
                self.log(f"CSRF: {csrf_token[:20]}...", Colors.SUCCESS)
                
                # === Step 4: POST signin → auth URL ===
                self.log("Redirecting to auth.openai.com...", Colors.INFO, "🔄 ")
                auth_session_id = str(uuid.uuid4())
                signin_js = """
                    async ([deviceId, authSessionId, loginEmail, csrfTok]) => {
                        const params = new URLSearchParams({
                            'prompt': 'login',
                            'ext-oai-did': deviceId,
                            'auth_session_logging_id': authSessionId,
                            'screen_hint': 'login_or_signup',
                            'login_hint': loginEmail
                        });
                        const r = await fetch('/api/auth/signin/openai?' + params.toString(), {
                            method: 'POST',
                            headers: { 'Content-Type': 'application/x-www-form-urlencoded' },
                            body: new URLSearchParams({
                                'callbackUrl': 'https://chatgpt.com/',
                                'csrfToken': csrfTok,
                                'json': 'true'
                            }).toString()
                        });
                        return await r.json();
                    }
                """
                signin_resp = page.evaluate(signin_js, [device_id, auth_session_id, email, csrf_token])
                auth_url = signin_resp.get("url", "")
                if not auth_url:
                    self.log("No auth URL returned", Colors.ERROR, "❌ ")
                    self.cleanup_browser()
                    self.current_retry += 1
                    continue
                self.log("Auth URL OK", Colors.SUCCESS)
            
                # Navigate to auth page
                if self.stop_event and self.stop_event.is_set():
                    self.cleanup_browser()
                    return (False, None)
                page.goto(auth_url, wait_until="domcontentloaded", timeout=60000)
                if self.stop_event and self.stop_event.is_set():
                    self.cleanup_browser()
                    return (False, None)
                page.wait_for_timeout(2000)
                
                # === Step 5: Register via API ===
                self.log("Registering account...", Colors.INFO, "📝 ")
                register_js = """
                    async ([regEmail, regPass]) => {
                        const r = await fetch('/api/accounts/user/register', {
                            method: 'POST',
                            headers: { 'Content-Type': 'application/json', 'Accept': 'application/json' },
                            body: JSON.stringify({ 'username': regEmail, 'password': regPass })
                        });
                        return { status: r.status, body: await r.text() };
                    }
                """
                register_resp = page.evaluate(register_js, [email, password])
                reg_status = register_resp.get('status', 0)
                reg_body = register_resp.get('body', '')
                self.log(f"Register: {reg_status}", Colors.SUCCESS if reg_status == 200 else Colors.ERROR)
                
                if reg_status != 200:
                    self.log(f"Register failed ({reg_status})", Colors.ERROR, "❌ ")
                    self.cleanup_browser()
                    self.current_retry += 1
                    continue
                
                # Follow continue_url to trigger OTP email
                try:
                    reg_data = json.loads(reg_body)
                    continue_url = reg_data.get('continue_url', '')
                    if continue_url:
                        self.log("Following continue_url...", Colors.INFO)
                        otp_trigger_js = """
                            async (url) => {
                                const r = await fetch(url, { headers: { 'Accept': 'application/json' } });
                                return { status: r.status, body: await r.text() };
                            }
                        """
                        page.evaluate(otp_trigger_js, continue_url)
                except Exception as e:
                    self.log(f"OTP trigger error: {e}", Colors.WARNING)
                
                if self.stop_event and self.stop_event.is_set():
                    self.cleanup_browser()
                    return (False, None)
                page.wait_for_timeout(2000)
                
                # === Step 6: Wait for OTP ===
                self.log("Waiting for OTP email...", Colors.INFO, "📬 ")
                if self.email_mode == "OAuth2":
                    otp_code = self._wait_for_otp_oauth2(timeout=120)
                else:
                    otp_code = self._wait_for_otp_tinyhost(email, timeout=120)
                
                if not otp_code:
                    self.cleanup_browser()
                    return (False, None)
                
                self.log(f"OTP: {otp_code}", Colors.SUCCESS, "✅ ")
                
                # Validate OTP via API
                validate_js = """
                    async (code) => {
                        const r = await fetch('/api/accounts/email-otp/validate', {
                            method: 'POST',
                            headers: { 'Content-Type': 'application/json', 'Accept': 'application/json' },
                            body: JSON.stringify({ 'code': code })
                        });
                        return { status: r.status, body: await r.text() };
                    }
                """
                validate_resp = page.evaluate(validate_js, otp_code)
                status = validate_resp.get('status')
                body = validate_resp.get('body', '')
                self.log(f"OTP Validate: {status}", Colors.SUCCESS if status == 200 else Colors.ERROR)
                if self.stop_event and self.stop_event.is_set():
                    self.cleanup_browser()
                    return (False, None)
                page.wait_for_timeout(2000)
                
                # === Step 7: Create account (name + DOB) ===
                self.log("Creating account...", Colors.INFO, "🏗️ ")
                create_js = """
                    async ([name, dob]) => {
                        const r = await fetch('/api/accounts/create_account', {
                            method: 'POST',
                            headers: { 'Content-Type': 'application/json', 'Accept': 'application/json' },
                            body: JSON.stringify({ 'name': name, 'birthdate': dob })
                        });
                        return { status: r.status, body: await r.text() };
                    }
                """
                create_resp = page.evaluate(create_js, [full_name, birthdate])
                self.log(f"Account created: {create_resp.get('status')}", Colors.SUCCESS if create_resp.get('status') == 200 else Colors.ERROR)
                if self.stop_event and self.stop_event.is_set():
                    self.cleanup_browser()
                    return (False, None)
                page.wait_for_timeout(2000)
                
                # Follow callback URL to complete login
                try:
                    create_data = json.loads(create_resp.get("body", "{}"))
                    continue_url = create_data.get("continue_url", "")
                    if continue_url and "callback" in continue_url:
                        self.log("Following auth callback...", Colors.INFO)
                        page.goto(continue_url, wait_until="domcontentloaded", timeout=60000)
                        page.wait_for_timeout(3000)
                except Exception as e:
                    self.log(f"Callback error: {e}", Colors.WARNING)
                
                if self.stop_event and self.stop_event.is_set():
                    self.cleanup_browser()
                    return (False, None)
                
                # === Verify registration ===
                self.log("Verifying registration...", Colors.INFO, "🔍 ")
                page.goto("https://chatgpt.com/", wait_until="domcontentloaded", timeout=60000)
                page.wait_for_timeout(2000)
                
                check_resp = page.evaluate("""
                    async () => {
                        const r = await fetch('/backend-api/accounts/check/v4-2023-04-27?timezone_offset_min=-420');
                        return { status: r.status };
                    }
                """)
                
                if check_resp.get("status") != 200:
                    self.log("Registration verification failed", Colors.ERROR, "❌ ")
                    self.cleanup_browser()
                    self.current_retry += 1
                    continue
                
                self.log("Registration verified!", Colors.SUCCESS, "✅ ")
                
                # === Get access token ===
                session_resp = page.evaluate("""
                    async () => {
                        const r = await fetch('/api/auth/session');
                        return await r.json();
                    }
                """)
                access_token = session_resp.get("accessToken", "")
                session_json_str = json.dumps(session_resp) if isinstance(session_resp, dict) else ""
                
                if access_token:
                    self.log(f"Access Token: {access_token[:30]}...", Colors.SUCCESS)
                
                if self.stop_event and self.stop_event.is_set():
                    # Save what we have so far before stopping
                    if access_token:
                        self.save_account_info(session_json_str)
                    self.cleanup_browser()
                    return (True if access_token else False, {'email': email, 'password': password} if access_token else None)
                
                # === 2FA setup if enabled ===
                mfa_secret = None
                if ENABLE_2FA and access_token:
                    mfa_result = self._setup_2fa_api(access_token)
                    if mfa_result:
                        mfa_secret = mfa_result.get("secret", "")
                
                # === Checkout if enabled ===
                checkout_url = None
                business_checkout_url = None
                if GET_CHECKOUT_LINK and access_token:
                    checkout_type = GET_CHECKOUT_TYPE
                    if checkout_type in ("Plus", "Both"):
                        checkout_url = self.get_checkout_link_via_api(access_token, "Plus")
                    if checkout_type in ("Business", "Both"):
                        business_checkout_url = self.get_checkout_link_via_api(access_token, "Business")
                
                # === Save to Excel ===
                saved = self.save_account_info(session_json_str, checkout_url, business_checkout_url, mfa_secret)
                
                if not saved:
                    self.log("Account not saved", Colors.ERROR, "❌ ")
                    self.cleanup_browser()
                    return (False, None)
                
                # Mark OAuth2 as registered if applicable
                if self.email_mode == "OAuth2" and self.oauth2_row_num:
                    mark_oauth2_registered(self.oauth2_row_num)
                
                self.log("─" * 40, Colors.SUCCESS)
                self.log("COMPLETED!", Colors.SUCCESS, "🎉 ")
                self.log(f"{email}:{password}", Colors.INFO)
                if mfa_secret:
                    self.log(f"2FA: {mfa_secret}", Colors.INFO, "🔐 ")
                self.log("─" * 40, Colors.SUCCESS)
                
                time.sleep(2)
                self.cleanup_browser()
                return (True, {'email': email, 'password': password})
                
            except Exception as e:
                # If stopping, exit immediately without retry/traceback
                if self.stop_event and self.stop_event.is_set():
                    self.cleanup_browser()
                    return (False, None)
                self.log(f"Error: {e}", Colors.ERROR, "❌ ")
                traceback.print_exc()
                self.cleanup_browser()
                self.current_retry += 1
                continue
        
        self.log(f"Failed after {self.max_retries} retries", Colors.ERROR, "❌ ")
        self.cleanup_browser()
        return (False, None)


def run_worker(thread_id, stop_event=None, thread_delay=2, num_threads=1, email_mode="TinyHost", oauth2_account=None):
    """Worker function for registration with staggered start and retry logic"""
    position_in_batch = (thread_id - 1) % num_threads if num_threads > 1 else 0
    delay = position_in_batch * thread_delay
    if delay > 0:
        safe_print(thread_id, f"Waiting {delay}s before starting...", Colors.INFO, "⏳ ")
        for _ in range(int(delay * 2)):
            if stop_event and stop_event.is_set():
                return (False, None)
            time.sleep(0.5)
    
    worker = ChatGPTAutoRegisterWorker(
        thread_id, 
        num_threads=num_threads,
        email_mode=email_mode,
        oauth2_account=oauth2_account
    )
    worker.stop_event = stop_event
    return worker.run()



# ============================================================================
# MODULE 4: CHECKOUT CAPTURE
# ============================================================================

# Shared TLS fingerprint list for checkout API calls
_TLS_FINGERPRINTS = []
for _ver in [110, 112, 114, 116, 118, 119, 120]:
    for _os_val in ["Windows NT 10.0; Win64; x64", "Windows NT 11.0; Win64; x64",
                     "Macintosh; Intel Mac OS X 10_15_7", "Macintosh; Intel Mac OS X 13_3"]:
        _TLS_FINGERPRINTS.append({"id": f"chrome_{_ver}", "ua": f"Mozilla/5.0 ({_os_val}) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/{_ver}.0.0.0 Safari/537.36"})
for _ver in [108, 110, 117, 120]:
    for _os_val in ["Windows NT 10.0; Win64; x64", "Windows NT 11.0; Win64; x64",
                     "Macintosh; Intel Mac OS X 10_15_7", "Macintosh; Intel Mac OS X 13_3"]:
        _TLS_FINGERPRINTS.append({"id": f"firefox_{_ver}", "ua": f"Mozilla/5.0 ({_os_val}; rv:{_ver}.0) Gecko/20100101 Firefox/{_ver}.0"})


def call_checkout_api(access_token, payload, label="Checkout", log_func=None):
    """Shared: POST /backend-api/payments/checkout → checkout URL.
    log_func(msg, color, emoji) is optional for logging.
    """
    fp = random.choice(_TLS_FINGERPRINTS)
    session = tls_client.Session(client_identifier=fp["id"], random_tls_extension_order=True)
    lang = random.choice(["en-US,en;q=0.9", "vi-VN,vi;q=0.9,en-US;q=0.8", "en-GB,en;q=0.9"])

    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json",
        "User-Agent": fp["ua"],
        "Accept-Language": lang,
        "Referer": "https://chatgpt.com/",
        "Origin": "https://chatgpt.com"
    }

    time.sleep(random.uniform(0.5, 1.5))
    resp = session.post("https://chatgpt.com/backend-api/payments/checkout", headers=headers, json=payload)

    if resp.status_code != 200:
        if log_func:
            log_func(f"{label} API error: {resp.status_code} - {resp.text[:80]}", Colors.ERROR, "❌ ")
        return None

    try:
        res_json = resp.json()
    except Exception:
        if log_func:
            log_func(f"{label} response not JSON", Colors.ERROR, "❌ ")
        return None

    final_link = res_json.get("url")
    if not final_link and res_json.get("checkout_session_id"):
        sid = res_json.get("checkout_session_id")
        final_link = f"https://chatgpt.com/checkout/openai_llc/{sid}"

    return final_link


class CheckoutCaptureWorker:
    """Worker for capturing checkout links from existing accounts"""
    
    def __init__(self, thread_id, email, access_token, excel_file, row_index, checkout_type="Plus"):
        self.thread_id = thread_id
        self.email = email
        self.access_token = access_token
        self.excel_file = excel_file
        self.row_index = row_index
        self.checkout_type = checkout_type  # "Plus", "Business", or "Both"
        
    def log(self, message, color=Colors.INFO, emoji=""):
        safe_print(self.thread_id, message, color, emoji)

    def save_to_excel(self, plus_url=None, business_url=None):
        """Save checkout URLs to Excel"""
        try:
            with file_lock:
                wb = load_workbook(self.excel_file)
                ws = wb.active
                
                if plus_url:
                    ws.cell(row=self.row_index, column=3, value=plus_url)
                    self.log("Saved Plus URL to Excel", Colors.SUCCESS, "💾 ")
                    
                if business_url:
                    ws.cell(row=self.row_index, column=4, value=business_url)
                    self.log("Saved Business URL to Excel", Colors.SUCCESS, "💾 ")
                
                wb.save(self.excel_file)
                wb.close()
                
            return True
        except Exception as e:
            self.log(f"Failed to save to Excel: {e}", Colors.ERROR, "❌ ")
            return False
    
    def save_no_plus_offer(self):
        """Save 'no Plus offer' to Excel column C when account has no free Plus offer"""
        try:
            with file_lock:
                wb = load_workbook(self.excel_file)
                ws = wb.active
                ws.cell(row=self.row_index, column=3, value="no Plus offer")
                wb.save(self.excel_file)
                wb.close()
                self.log("Saved 'no Plus offer' to Excel", Colors.SUCCESS, "💾 ")
            return True
        except Exception as e:
            self.log(f"Failed to save no Plus offer: {e}", Colors.ERROR, "❌ ")
            return False
    
    def run(self):
        """Run the checkout capture flow via API"""
        try:
            self.log(f"Checkout capture for {self.email}", Colors.HEADER, "🚀 ")

            access_token = self.access_token
            if not access_token:
                self.log("No access token", Colors.ERROR, "❌ ")
                return False

            self.log(f"Using access token: {access_token[:30]}...", Colors.SUCCESS, "🔑 ")

            plus_url = None
            business_url = None

            if self.checkout_type in ["Plus", "Both"]:
                self.log("Requesting Plus checkout link...", Colors.INFO, "💳 ")
                plus_url = call_checkout_api(access_token, ChatGPTAutoRegisterWorker.CHECKOUT_PAYLOADS["Plus"], "Plus", self.log)
                if plus_url:
                    self.log(f"Plus URL: {plus_url[:60]}...", Colors.SUCCESS, "✅ ")
                else:
                    self.log("Could not get Plus link (no offer or error)", Colors.WARNING, "⚠️ ")
                    self.save_no_plus_offer()

            if self.checkout_type in ["Business", "Both"]:
                self.log("Requesting Business checkout link...", Colors.INFO, "💼 ")
                business_url = call_checkout_api(access_token, ChatGPTAutoRegisterWorker.CHECKOUT_PAYLOADS["Business"], "Business", self.log)
                if business_url:
                    self.log(f"Business URL: {business_url[:60]}...", Colors.SUCCESS, "✅ ")
                else:
                    self.log("Could not get Business link", Colors.WARNING, "⚠️ ")

            # Save results
            if plus_url or business_url:
                self.save_to_excel(plus_url, business_url)
                self.log(f"✅ Completed for {self.email}", Colors.SUCCESS, "🎉 ")
                return True
            else:
                self.log(f"No checkout URLs captured for {self.email}", Colors.WARNING, "⚠️ ")
                return False

        except Exception as e:
            self.log(f"Error: {e}", Colors.ERROR, "❌ ")
            return False


def load_checkout_accounts(excel_file):
    """Load all accounts from Excel for checkout capture"""
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        
        accounts = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if len(row) < 2:
                continue
            
            account = row[0] if row[0] else ""
            
            # Smart detect: JSON format (new) vs plain accessToken (legacy)
            raw_col_b = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            access_token = ""
            if raw_col_b:
                if raw_col_b.startswith("{"):
                    try:
                        parsed = json.loads(raw_col_b)
                        access_token = parsed.get("accessToken", "")
                    except Exception:
                        access_token = raw_col_b
                else:
                    access_token = raw_col_b  # Legacy plain token format
            
            plus_url = row[2] if len(row) > 2 and row[2] else ""
            business_url = row[3] if len(row) > 3 and row[3] else ""
            sold_status = row[5] if len(row) > 5 and row[5] else ""
            
            if not account or not access_token:
                continue
            
            # Parse email from account
            if ":" in str(account):
                email = str(account).split(":", 1)[0]
            else:
                email = str(account)
            
            # Check if sold (any value in column F means sold)
            is_sold = bool(sold_status and str(sold_status).strip())
            
            accounts.append({
                "row_index": row_idx,
                "email": email,
                "account": account,
                "access_token": access_token,
                "plus_url": plus_url,
                "business_url": business_url,
                "sold_status": sold_status,
                "is_sold": is_sold,
            })
        
        wb.close()
        return accounts
        
    except Exception as e:
        print(f"Error loading accounts: {e}")
        return []


# ============================================================================
# GUI IMPLEMENTATION
# ============================================================================

# Settings
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

# ============================================================================
# ENHANCED UI UTILITIES
# ============================================================================

class GlowButton(ctk.CTkButton):
    """Custom button with glow effect on hover"""
    def __init__(self, master, glow_color="#00d4ff", **kwargs):
        super().__init__(master, **kwargs)
        self.glow_color = glow_color
        self.default_border = kwargs.get("border_width", 0)
        self.bind("<Enter>", self._on_hover_enter)
        self.bind("<Leave>", self._on_hover_leave)
        
    def configure(self, require_redraw=False, **kwargs):
        if "glow_color" in kwargs:
            self.glow_color = kwargs.pop("glow_color")
        super().configure(require_redraw=require_redraw, **kwargs)

    def _on_hover_enter(self, e=None):
        try:
            self.configure(border_width=2, border_color=self.glow_color)
        except Exception: pass
        
    def _on_hover_leave(self, e=None):
        try:
            self.configure(border_width=self.default_border, border_color="transparent")
        except Exception: pass


class AnimatedCard(ctk.CTkFrame):
    """Card with entrance animation"""
    def __init__(self, master, delay_ms=0, **kwargs):
        # Default styling
        kwargs.setdefault("corner_radius", 16)
        kwargs.setdefault("border_width", 1)
        kwargs.setdefault("border_color", "#30363d")
        super().__init__(master, **kwargs)
        self.delay_ms = delay_ms
        self._initial_alpha = 0
        
    def animate_in(self, motion):
        """Trigger entrance animation"""
        def do_animate():
            motion.color(self, "fg_color", self.cget("fg_color"), duration_ms=400, steps=30)
        self.after(self.delay_ms, do_animate)


class PulsingDot(ctk.CTkFrame):
    """Animated status dot"""
    def __init__(self, master, color="#00ff88", size=12, **kwargs):
        super().__init__(master, width=size, height=size, corner_radius=size//2, fg_color=color, **kwargs)
        self.base_color = color
        self.is_pulsing = False
        self._pulse_job = None
        
    def start_pulse(self):
        if self.is_pulsing:
            return
        self.is_pulsing = True
        self._do_pulse(True)
        
    def _do_pulse(self, bright):
        if not self.is_pulsing:
            return
        try:
            target = self.base_color if bright else self._dim_color(self.base_color)
            self.configure(fg_color=target)
            self._pulse_job = self.after(500, lambda: self._do_pulse(not bright))
        except Exception: pass
        
    def _dim_color(self, hex_color):
        """Dim a hex color by 50%"""
        h = hex_color.lstrip("#")
        r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
        return f"#{r//2:02x}{g//2:02x}{b//2:02x}"
        
    def stop_pulse(self):
        self.is_pulsing = False
        if self._pulse_job:
            try:
                self.after_cancel(self._pulse_job)
            except Exception: pass
        self.configure(fg_color=self.base_color)


class TextRedirector(object):
    def __init__(self, widget, tag="stdout"):
        self.widget = widget
        self.tag = tag
        # ANSI codes map
        self.tag_map = {
            '\033[31m': 'error',     # Red
            '\033[32m': 'success',   # Green
            '\033[33m': 'warning',   # Yellow
            '\033[34m': 'info',      # Blue
            '\033[36m': 'info',      # Cyan (mapped to info)
            '\033[35m': 'header',    # Magenta
            '\033[0m': 'normal',     # Reset
            '\x1b[31m': 'error',
            '\x1b[32m': 'success',
            '\x1b[33m': 'warning',
            '\x1b[34m': 'info',
            '\x1b[36m': 'info',
            '\x1b[35m': 'header',
            '\x1b[0m': 'normal',
            '\x1b[39m': 'normal'
        }
        self.ansi_re = re.compile(r'(\x1B\[[0-9;]*m)')

    # Class-level flag to suppress output during stop (bulletproof, no widget traversal)
    _suppress_output = False

    def write(self, str_data):
        if str_data and not TextRedirector._suppress_output:
            self.widget.after(0, self._append_text, str_data)

    def _append_text(self, text):
        # Also check here to discard callbacks that were already queued before suppress
        if TextRedirector._suppress_output:
            return
        try:
            self.widget.configure(state="normal")
            parts = self.ansi_re.split(text)
            current_tag = "normal"
            
            for part in parts:
                if part in self.tag_map:
                    current_tag = self.tag_map[part]
                elif part.startswith('\x1B'):
                    pass 
                elif part:
                    self.widget.insert("end", part, current_tag)
            
            self.widget.see("end")
            self.widget.configure(state="disabled")
        except Exception:
            pass

    def flush(self):
        pass

# =========================
# UI MOTION SYSTEM (CTk)
# =========================
import math

class MotionTokens:
    # Durations (ms)
    fast = 120
    normal = 180
    slow = 260
    pulse_period = 950

    # Easing
    @staticmethod
    def ease_out_quad(t: float) -> float:
        return 1.0 - (1.0 - t) * (1.0 - t)

    @staticmethod
    def ease_out_cubic(t: float) -> float:
        return 1.0 - (1.0 - t) ** 3

def _hex_to_rgb(h: str):
    h = h.strip()
    if h.startswith("#"): h = h[1:]
    if len(h) == 3: h = "".join([c * 2 for c in h])
    return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)

def _rgb_to_hex(rgb):
    return f"#{rgb[0]:02x}{rgb[1]:02x}{rgb[2]:02x}"

def _lerp(a, b, t): 
    return a + (b - a) * t

class Motion:
    """
    Motion engine cho CustomTkinter:
    - cancel theo key
    - color tween
    - number tween
    - pulse loop
    - hover transitions
    """
    def __init__(self, app, tokens=MotionTokens):
        self.app = app
        self.tok = tokens
        self._jobs = {}          # key -> after job id
        self._pulse_jobs = {}    # key -> after job id
        self._pulse_on = set()   # keys đang pulse

    def _key(self, widget, prop, suffix=""):
        return f"{id(widget)}::{prop}::{suffix}"

    def cancel(self, key: str):
        job = self._jobs.pop(key, None)
        if job is not None:
            try: self.app.after_cancel(job)
            except Exception: pass

    def color(self, widget, prop: str, to_hex: str, *,
              duration_ms=None, steps=18, easing=None):
        duration_ms = duration_ms if duration_ms is not None else self.tok.normal
        easing = easing if easing is not None else self.tok.ease_out_quad

        try:
            cur = widget.cget(prop)
        except Exception:
            try:
                widget.configure(**{prop: to_hex})
                return
            except Exception:
                return

        if not isinstance(cur, str) or not cur.startswith("#") or len(cur) not in (4, 7):
            try:
                widget.configure(**{prop: to_hex})
            except Exception:
                pass
            return

        try:
            sr, sg, sb = _hex_to_rgb(cur)
            er, eg, eb = _hex_to_rgb(to_hex)
        except Exception:
            try: widget.configure(**{prop: to_hex})
            except Exception: pass
            return

        key = self._key(widget, prop, "color")
        self.cancel(key)

        start_time = time.time()

        def tick():
            t = (time.time() - start_time) * 1000.0 / max(duration_ms, 1)
            t = max(0.0, min(1.0, t))
            te = easing(t)

            r = int(_lerp(sr, er, te))
            g = int(_lerp(sg, eg, te))
            b = int(_lerp(sb, eb, te))

            try:
                widget.configure(**{prop: _rgb_to_hex((r, g, b))})
            except Exception:
                pass

            if t < 1.0:
                job = self.app.after(max(1, duration_ms // steps), tick)
                self._jobs[key] = job
            else:
                self._jobs.pop(key, None)

        tick()

    def number(self, setter, start: float, end: float, *,
               duration_ms=None, steps=18, easing=None, fmt=None):
        duration_ms = duration_ms if duration_ms is not None else self.tok.normal
        easing = easing if easing is not None else self.tok.ease_out_quad
        key = f"{id(setter)}::number"

        self.cancel(key)
        start_time = time.time()

        def tick():
            t = (time.time() - start_time) * 1000.0 / max(duration_ms, 1)
            t = max(0.0, min(1.0, t))
            te = easing(t)
            val = _lerp(start, end, te)
            try:
                setter(fmt(val) if fmt else val)
            except Exception:
                pass

            if t < 1.0:
                job = self.app.after(max(1, duration_ms // steps), tick)
                self._jobs[key] = job
            else:
                self._jobs.pop(key, None)

        tick()

    def pulse(self, widget, prop: str, a: str, b: str, *,
              period_ms=None):
        period_ms = period_ms if period_ms is not None else self.tok.pulse_period
        key = self._key(widget, prop, "pulse")

        self._pulse_on.add(key)

        job = self._pulse_jobs.pop(key, None)
        if job is not None:
            try: self.app.after_cancel(job)
            except Exception: pass

        def loop(state=0):
            if key not in self._pulse_on:
                return
            start, end = (a, b) if state == 0 else (b, a)
            self.color(widget, prop, end, duration_ms=period_ms // 2, steps=20, easing=self.tok.ease_out_quad)
            job2 = self.app.after(period_ms // 2, lambda: loop(1 - state))
            self._pulse_jobs[key] = job2

        loop(0)

    def stop_pulse(self, widget, prop: str):
        key = self._key(widget, prop, "pulse")
        if key in self._pulse_on:
            self._pulse_on.remove(key)
        job = self._pulse_jobs.pop(key, None)
        if job is not None:
            try: self.app.after_cancel(job)
            except Exception: pass

    def hover(self, widget, *,
              enter=None, leave=None,
              duration_ms=None):
        duration_ms = duration_ms if duration_ms is not None else self.tok.fast
        enter = enter or {}
        leave = leave or {}

        def on_enter(_):
            for prop, val in enter.items():
                self.color(widget, prop, val, duration_ms=duration_ms)

        def on_leave(_):
            for prop, val in leave.items():
                self.color(widget, prop, val, duration_ms=duration_ms)

        try:
            widget.bind("<Enter>", on_enter)
            widget.bind("<Leave>", on_leave)
        except Exception:
            pass

def _toast(app, message, duration_ms=2500, toast_type="info"):
    """Modern animated toast notification"""
    try:
        toast = ctk.CTkToplevel(app)
        toast.overrideredirect(True)
        toast.attributes("-topmost", True)
        toast.attributes("-alpha", 0.0)  # Start invisible
        
        app.update_idletasks()
        w, h = 320, 56
        x = app.winfo_x() + app.winfo_width() - w - 24
        y = app.winfo_y() + 50
        toast.geometry(f"{w}x{h}+{x}+{y}")
        
        # Color based on type
        colors = {
            "info": ("#0f172a", "#00f0ff", "#1a3a4a"),
            "success": ("#0f172a", "#10b981", "#1a3d32"),
            "error": ("#0f172a", "#ef4444", "#3d1a1a"),
            "warning": ("#0f172a", "#f59e0b", "#3d2e1a"),
        }
        bg, accent, border = colors.get(toast_type, colors["info"])
        
        frame = ctk.CTkFrame(
            toast, 
            corner_radius=14, 
            fg_color=bg, 
            border_width=1, 
            border_color=border
        )
        frame.pack(fill="both", expand=True, padx=2, pady=2)
        
        # Accent bar
        accent_bar = ctk.CTkFrame(frame, width=4, height=40, corner_radius=2, fg_color=accent)
        accent_bar.place(x=12, rely=0.5, anchor="w")
        
        ctk.CTkLabel(
            frame, 
            text=message, 
            font=app._font("Segoe UI", 13, "bold"),
            text_color="#f8fafc"
        ).pack(expand=True, padx=(28, 16))
        
        # Fade in animation
        def fade_in(alpha=0.0):
            if alpha < 0.95:
                toast.attributes("-alpha", alpha)
                toast.after(16, lambda: fade_in(alpha + 0.1))
            else:
                toast.attributes("-alpha", 0.95)
        
        # Fade out animation
        def fade_out(alpha=0.95):
            if alpha > 0.05:
                toast.attributes("-alpha", alpha)
                toast.after(16, lambda: fade_out(alpha - 0.08))
            else:
                toast.destroy()
        
        fade_in()
        toast.after(duration_ms, fade_out)
        
    except Exception: pass

class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        # UI Motion Engine
        self.motion = Motion(self, MotionTokens)
        
        # State tracking
        self.process_running = False
        self._is_stopping = False  # Suppress logs during stop to prevent GUI flood

        # Window setup
        self.title("⚡ ChatGPT Auto Tools")
        self.geometry("1100x920")
        self.minsize(900, 750)
        
        # Configure window background
        self.configure(fg_color="#080b12")
        
        # Premium Fonts - More distinctive choices
        self._font_cache = {}  # Memoize CTkFont to avoid creating duplicate objects
        self.font_title = self._font("Segoe UI", 36, "bold")
        self.font_subtitle = self._font("Segoe UI Light", 15)
        self.font_label = self._font("Segoe UI", 13)
        self.font_button = self._font("Segoe UI Semibold", 14, "bold")
        self.font_mono = self._font("Cascadia Code", 11)
        self.font_stats = self._font("Segoe UI", 28, "bold")
        self.font_stats_label = self._font("Segoe UI", 11)
        
        # 🎨 VIBRANT Cyberpunk Color Palette
        self.colors = {
            # Primary Accents - Electric Neons
            "accent_primary": "#00f0ff",    # Electric Cyan
            "accent_secondary": "#ff00aa",  # Hot Pink/Magenta
            "accent_tertiary": "#7c3aed",   # Purple
            "accent_purple": "#a855f7",     # Vivid Purple (NEW)
            "accent_cyan": "#06b6d4",       # Cyan (NEW)
            "accent_green": "#00ff9f",      # Neon Mint
            "accent_orange": "#ff6b35",     # Coral Orange
            "accent_yellow": "#ffd600",     # Electric Yellow
            
            # Backgrounds - Deep Space
            "bg_base": "#080b12",           # Near Black
            "bg_dark": "#0c1018",           # Dark Navy
            "bg_card": "#111827",           # Card Background
            "bg_card_hover": "#1a2332",     # Card Hover
            "bg_elevated": "#1e293b",       # Elevated surfaces
            
            # Glass Effects
            "glass_bg": "#0f172a",          # Glassmorphism base
            "glass_border": "#334155",      # Glass border
            
            # Gradients (start colors)
            "gradient_cyan": "#00f0ff",
            "gradient_purple": "#a855f7",
            "gradient_pink": "#ec4899",
            
            # Text
            "text_primary": "#f8fafc",      # Bright white
            "text_secondary": "#94a3b8",    # Slate gray
            "text_muted": "#64748b",        # Muted
            
            # Status Colors
            "success": "#10b981",           # Emerald
            "success_glow": "#34d399",      # Light emerald
            "error": "#ef4444",             # Red
            "error_glow": "#f87171",        # Light red
            "warning": "#f59e0b",           # Amber
            "info": "#3b82f6",              # Blue
            
            # Borders
            "border_subtle": "#1e293b",
            "border_glow": "#334155",
            "border_accent": "#1a3a4a",   # Cyan tint (muted)
        }

        # Grid layout (2 cols: Controls | Status)
        self.grid_columnconfigure(0, weight=3)  # Left side (Controls) - larger
        self.grid_columnconfigure(1, weight=2)  # Right side (Status)
        self.grid_rowconfigure(0, weight=0)     # Header
        self.grid_rowconfigure(1, weight=0)     # Main Content
        self.grid_rowconfigure(2, weight=1)     # Logs
        
        # --- SIMPLE HEADER ---
        self.header_label = ctk.CTkLabel(
            self, text="⚡ ChatGPT Auto Tools",
            font=self._font("Segoe UI", 22, "bold"),
            text_color=self.colors["accent_primary"]
        )
        self.header_label.grid(row=0, column=0, columnspan=2, padx=24, pady=(12, 4), sticky="w")

        # --- LEFT COLUMN: CONTROLS (ENHANCED TABVIEW) ---
        self.controls_container = ctk.CTkFrame(
            self,
            fg_color=self.colors["glass_bg"],
            corner_radius=20,
            border_width=1,
            border_color=self.colors["glass_border"]
        )
        self.controls_container.grid(row=1, column=0, padx=(32, 12), pady=12, sticky="nsew")
        
        self.tabview = ctk.CTkTabview(
            self.controls_container, 
            width=600,
            fg_color="transparent",
            segmented_button_fg_color=self.colors["bg_card"],
            segmented_button_selected_color=self.colors["accent_primary"],
            segmented_button_selected_hover_color=self.colors["accent_primary"],
            segmented_button_unselected_color=self.colors["bg_card"],
            segmented_button_unselected_hover_color=self.colors["bg_card_hover"],
            text_color=self.colors["text_primary"],
            corner_radius=16
        )
        self.tabview.pack(fill="both", expand=True, padx=16, pady=16)
        self.tabview.add("🚀 Registration")
        self.tabview.add("💳 Checkout Capture")
        
        # Enhanced tab font
        self.tabview._segmented_button.configure(
            font=self._font("Segoe UI Semibold", 13, "bold"),
            corner_radius=12
        )

        self.setup_registration_tab()
        self.setup_checkout_tab()
        
        # --- RIGHT COLUMN: STATUS PANEL (GLASSMORPHISM STYLE) ---
        self.status_frame = ctk.CTkFrame(
            self, 
            fg_color=self.colors["glass_bg"],
            corner_radius=20,
            border_width=1,
            border_color=self.colors["glass_border"]
        )
        self.status_frame.grid(row=1, column=1, padx=(12, 32), pady=12, sticky="nsew")
        self.status_frame.grid_columnconfigure(0, weight=1)
        
        # ═══ STATUS SECTION ═══
        self.status_header = ctk.CTkFrame(self.status_frame, fg_color="transparent")
        self.status_header.grid(row=0, column=0, padx=20, pady=(20, 12), sticky="ew")
        self.status_header.grid_columnconfigure(1, weight=1)
        
        # Pulsing status dot
        self.status_dot = PulsingDot(self.status_header, color=self.colors["text_muted"], size=10)
        self.status_dot.grid(row=0, column=0, padx=(0, 10), sticky="w")
        
        self.status_label = ctk.CTkLabel(
            self.status_header, 
            text="SYSTEM STATUS", 
            font=self._font("Segoe UI", 10, "bold"),
            text_color=self.colors["text_muted"]
        )
        self.status_label.grid(row=0, column=1, sticky="w")
        
        # Status indicator pill
        self.status_indicator = ctk.CTkButton(
            self.status_frame, 
            text="● IDLE", 
            fg_color=self.colors["bg_elevated"], 
            state="disabled", 
            width=140, height=36, 
            font=self._font("Segoe UI", 12, "bold"), 
            corner_radius=18,
            border_width=1,
            border_color=self.colors["border_subtle"]
        )
        self.status_indicator.grid(row=1, column=0, padx=20, pady=(0, 16), sticky="w")
        
        # ═══ STATS CARDS ═══
        self.stats_container = ctk.CTkFrame(self.status_frame, fg_color="transparent")
        self.stats_container.grid(row=2, column=0, padx=20, pady=8, sticky="ew")
        self.stats_container.grid_columnconfigure(0, weight=1)
        self.stats_container.grid_columnconfigure(1, weight=1)
        
        # Success Card
        self.success_card = ctk.CTkFrame(
            self.stats_container,
            fg_color=self.colors["bg_card"],
            corner_radius=14,
            border_width=1,
            border_color="#1a3d32"  # Green tint
        )
        self.success_card.grid(row=0, column=0, padx=(0, 6), pady=4, sticky="ew")
        
        self.success_icon = ctk.CTkLabel(
            self.success_card,
            text="✓",
            font=self._font("Segoe UI", 18, "bold"),
            text_color=self.colors["success"]
        )
        self.success_icon.pack(pady=(12, 4))
        
        self.success_count_label = ctk.CTkLabel(
            self.success_card,
            text="0",
            font=self.font_stats,
            text_color=self.colors["success_glow"]
        )
        self.success_count_label.pack(pady=(0, 2))
        
        self.success_text_label = ctk.CTkLabel(
            self.success_card,
            text="SUCCESS",
            font=self.font_stats_label,
            text_color=self.colors["text_muted"]
        )
        self.success_text_label.pack(pady=(0, 12))
        
        # Failed Card
        self.fail_card = ctk.CTkFrame(
            self.stats_container,
            fg_color=self.colors["bg_card"],
            corner_radius=14,
            border_width=1,
            border_color="#3d1a1a"  # Red tint
        )
        self.fail_card.grid(row=0, column=1, padx=(6, 0), pady=4, sticky="ew")
        
        self.fail_icon = ctk.CTkLabel(
            self.fail_card,
            text="✗",
            font=self._font("Segoe UI", 18, "bold"),
            text_color=self.colors["error"]
        )
        self.fail_icon.pack(pady=(12, 4))
        
        self.fail_count_label = ctk.CTkLabel(
            self.fail_card,
            text="0",
            font=self.font_stats,
            text_color=self.colors["error_glow"]
        )
        self.fail_count_label.pack(pady=(0, 2))
        
        self.fail_text_label = ctk.CTkLabel(
            self.fail_card,
            text="FAILED",
            font=self.font_stats_label,
            text_color=self.colors["text_muted"]
        )
        self.fail_text_label.pack(pady=(0, 12))
        
        # ═══ PROGRESS SECTION ═══
        self.progress_label = ctk.CTkLabel(
            self.status_frame, 
            text="PROGRESS",
            font=self._font("Segoe UI", 10, "bold"),
            text_color=self.colors["text_muted"]
        )
        self.progress_label.grid(row=3, column=0, padx=20, pady=(20, 8), sticky="w")
        
        # Custom gradient-like progress bar container
        self.progress_container = ctk.CTkFrame(
            self.status_frame,
            fg_color=self.colors["bg_card"],
            corner_radius=8,
            height=20
        )
        self.progress_container.grid(row=4, column=0, padx=20, pady=(0, 8), sticky="ew")
        
        self.progress_bar = ctk.CTkProgressBar(
            self.progress_container, 
            orientation="horizontal", 
            mode="determinate",
            height=6,
            progress_color=self.colors["accent_primary"],
            fg_color=self.colors["bg_elevated"],
            corner_radius=3
        )
        self.progress_bar.pack(fill="x", padx=8, pady=7)
        self.progress_bar.set(0)
        
        # Progress percentage
        self.progress_percent = ctk.CTkLabel(
            self.status_frame,
            text="0%",
            font=self._font("Segoe UI", 11, "bold"),
            text_color=self.colors["accent_primary"]
        )
        self.progress_percent.grid(row=5, column=0, padx=20, pady=(0, 12), sticky="e")

        # ═══ ACTIVITY LOG (Mini) ═══
        self.activity_label = ctk.CTkLabel(
            self.status_frame,
            text="ACTIVITY",
            font=self._font("Segoe UI", 10, "bold"),
            text_color=self.colors["text_muted"]
        )
        self.activity_label.grid(row=6, column=0, padx=20, pady=(8, 8), sticky="w")
        
        self.info_box = ctk.CTkTextbox(
            self.status_frame, 
            height=80, 
            font=self.font_mono, 
            fg_color=self.colors["bg_card"], 
            text_color=self.colors["accent_primary"], 
            corner_radius=10,
            border_width=1,
            border_color=self.colors["border_subtle"]
        )
        self.info_box.grid(row=7, column=0, padx=20, pady=(0, 20), sticky="ew")
        self.info_box.insert("0.0", "⚡ Ready to start automation...")
        self.info_box.configure(state="disabled")

        # --- LOG CONSOLE (TERMINAL STYLE) ---
        self.log_frame = ctk.CTkFrame(
            self,
            fg_color=self.colors["glass_bg"],
            corner_radius=20,
            border_width=1,
            border_color=self.colors["glass_border"]
        )
        self.log_frame.grid(row=2, column=0, columnspan=2, padx=24, pady=(8, 16), sticky="nsew")
        self.log_frame.grid_columnconfigure(0, weight=1)
        self.log_frame.grid_rowconfigure(1, weight=1)
        
        # Terminal-style toolbar
        self.log_toolbar = ctk.CTkFrame(self.log_frame, fg_color="transparent", height=40)
        self.log_toolbar.grid(row=0, column=0, padx=16, pady=(16, 8), sticky="ew")
        
        # Terminal window dots (macOS style)
        self.dots_frame = ctk.CTkFrame(self.log_toolbar, fg_color="transparent")
        self.dots_frame.pack(side="left", padx=(4, 12))
        
        for color in ["#ff5f57", "#febc2e", "#28c840"]:
            dot = ctk.CTkFrame(self.dots_frame, width=12, height=12, corner_radius=6, fg_color=color)
            dot.pack(side="left", padx=2)
        
        self.log_title = ctk.CTkLabel(
            self.log_toolbar, 
            text="SYSTEM CONSOLE", 
            font=self._font("Segoe UI", 11, "bold"),
            text_color=self.colors["text_muted"]
        )
        self.log_title.pack(side="left", padx=8)
        
        # Terminal blinking cursor indicator
        self.cursor_indicator = ctk.CTkLabel(
            self.log_toolbar,
            text="▊",
            font=self._font("Segoe UI", 12),
            text_color=self.colors["accent_primary"]
        )
        self.cursor_indicator.pack(side="left", padx=4)
        self.after(2000, self._blink_cursor)  # Defer until window is rendered
        
        # Toolbar buttons with modern styling
        btn_style = {
            "width": 36, "height": 28,
            "fg_color": self.colors["bg_card"],
            "hover_color": self.colors["bg_card_hover"],
            "corner_radius": 8,
            "font": self._font("Segoe UI", 14),
            "text_color": self.colors["text_secondary"],
            "border_width": 1,
            "border_color": self.colors["border_subtle"]
        }
        
        self.btn_clear = ctk.CTkButton(self.log_toolbar, text="🗑", command=self.clear_logs, **btn_style)
        self.btn_clear.pack(side="right", padx=3)
        
        self.btn_copy = ctk.CTkButton(self.log_toolbar, text="📋", command=self.copy_logs, **btn_style)
        self.btn_copy.pack(side="right", padx=3)
        
        self.btn_export = ctk.CTkButton(self.log_toolbar, text="💾", command=self.export_logs, **btn_style)
        self.btn_export.pack(side="right", padx=3)

        # Log textbox with terminal aesthetics
        self.log_textbox = ctk.CTkTextbox(
            self.log_frame, 
            font=self.font_mono, 
            fg_color=self.colors["bg_base"],
            text_color="#e2e8f0",
            corner_radius=12,
            border_width=1,
            border_color=self.colors["border_subtle"],
            scrollbar_button_color=self.colors["bg_elevated"],
            scrollbar_button_hover_color=self.colors["accent_primary"]
        )
        self.log_textbox.grid(row=1, column=0, padx=16, pady=(0, 16), sticky="nsew")
        
        # Configure Tags for Colored Logs (Neon Cyberpunk)
        self.log_textbox.tag_config("error", foreground="#f87171")      # Light red
        self.log_textbox.tag_config("success", foreground="#4ade80")    # Light green
        self.log_textbox.tag_config("warning", foreground="#fbbf24")    # Amber
        self.log_textbox.tag_config("info", foreground="#38bdf8")       # Sky blue
        self.log_textbox.tag_config("header", foreground="#c084fc")     # Purple
        self.log_textbox.tag_config("normal", foreground="#e2e8f0")     # Slate
        
        self.log_textbox.configure(state="disabled")

        # Redirect stdout
        sys.stdout = TextRedirector(self.log_textbox)
        sys.stderr = TextRedirector(self.log_textbox)
        self.running = False
        self.stop_event = threading.Event()
        self.log_buffer = []
        self._active_workers = set()
        self._active_executors = set()
        self._active_runtime_lock = threading.Lock()
        
        pass  # UI ready

    # --- FONT CACHE ---
    def _font(self, family="Segoe UI", size=13, weight="normal", slant="roman"):
        """Memoized CTkFont — avoids creating duplicate font objects (53→~15 unique)"""
        key = (family, size, weight, slant)
        if key not in self._font_cache:
            self._font_cache[key] = ctk.CTkFont(family=family, size=size, weight=weight, slant=slant)
        return self._font_cache[key]

    # --- SETUP TABS ---
    def setup_registration_tab(self):
        tab = self.tabview.tab("🚀 Registration")
        tab.grid_columnconfigure(0, weight=1)
        
        # ═══ SETTINGS CARD ═══
        settings_card = ctk.CTkFrame(
            tab, 
            fg_color=self.colors["bg_card"],
            corner_radius=16,
            border_width=1,
            border_color=self.colors["border_subtle"]
        )
        settings_card.pack(fill="x", padx=8, pady=(4, 8))
        
        # Card Header
        card_header = ctk.CTkFrame(settings_card, fg_color="transparent")
        card_header.pack(fill="x", padx=16, pady=(6, 4))
        
        ctk.CTkLabel(
            card_header, 
            text="⚙️  Configuration", 
            font=self._font("Segoe UI Semibold", 14, "bold"),
            text_color=self.colors["text_primary"]
        ).pack(side="left")
        
        # Divider
        ctk.CTkFrame(settings_card, height=1, fg_color=self.colors["border_subtle"]).pack(fill="x", padx=16)
        
        # Mode Selection
        self.reg_mode_frame = ctk.CTkFrame(settings_card, fg_color="transparent")
        self.reg_mode_frame.pack(fill="x", padx=16, pady=6)
        
        mode_label = ctk.CTkLabel(
            self.reg_mode_frame, 
            text="Execution Mode", 
            font=self.font_label,
            text_color=self.colors["text_secondary"]
        )
        mode_label.pack(side="left", padx=(0, 16))
        
        self.reg_mode_var = ctk.StringVar(value="Sequential")
        self.reg_mode_menu = ctk.CTkOptionMenu(
            self.reg_mode_frame,
            values=["Sequential", "Multithread"],
            variable=self.reg_mode_var,
            command=self.toggle_reg_inputs,
            width=160,
            height=32,
            font=self.font_label,
            dropdown_font=self.font_label,
            fg_color=self.colors["accent_tertiary"],
            button_color=self.colors["accent_secondary"],
            button_hover_color=self.colors["accent_primary"],
            dropdown_fg_color=self.colors["bg_card"],
            dropdown_hover_color=self.colors["bg_card_hover"],
            corner_radius=10
        )
        self.reg_mode_menu.pack(side="right")
        
        # Count Input Row (will be modified for Multithread to include Threads)
        self.reg_count_frame = ctk.CTkFrame(settings_card, fg_color="transparent")
        self.reg_count_frame.pack(fill="x", padx=16, pady=(0, 6))
        
        # Left side: Total Accounts
        self.reg_count_left = ctk.CTkFrame(self.reg_count_frame, fg_color="transparent")
        self.reg_count_left.pack(side="left", fill="x", expand=True)
        
        self.reg_count_label = ctk.CTkLabel(
            self.reg_count_left, 
            text="Account Count", 
            font=self.font_label,
            text_color=self.colors["text_secondary"]
        )
        self.reg_count_label.pack(side="left", padx=(0, 8))
        
        self.reg_count_entry = ctk.CTkEntry(
            self.reg_count_left, 
            placeholder_text="1", 
            width=70,
            height=36,
            font=self.font_label,
            fg_color=self.colors["bg_elevated"],
            border_color=self.colors["border_subtle"],
            corner_radius=10
        )
        self.reg_count_entry.insert(0, "1")
        self.reg_count_entry.pack(side="left", padx=(0, 0))
        
        # Right side: Threads (only visible in Multithread mode)
        self.reg_threads_frame = ctk.CTkFrame(self.reg_count_frame, fg_color="transparent")
        # Hidden by default (Sequential mode)
        
        self.reg_threads_label = ctk.CTkLabel(
            self.reg_threads_frame, 
            text="Threads", 
            font=self.font_label,
            text_color=self.colors["text_secondary"]
        )
        self.reg_threads_label.pack(side="left", padx=(0, 8))
        
        self.reg_threads_entry = ctk.CTkEntry(
            self.reg_threads_frame, 
            placeholder_text="2", 
            width=70,
            height=36,
            font=self.font_label,
            fg_color=self.colors["bg_elevated"],
            border_color=self.colors["border_subtle"],
            corner_radius=10
        )
        self.reg_threads_entry.insert(0, "2")
        self.reg_threads_entry.pack(side="left")
        
        # Thread Delay Input (only visible in Multithread mode)
        self.reg_delay_frame = ctk.CTkFrame(settings_card, fg_color="transparent")
        self.reg_delay_frame.pack(fill="x", padx=16, pady=(0, 10))
        self.reg_delay_frame.pack_forget()  # Hidden by default (Sequential mode)
        
        self.reg_delay_label = ctk.CTkLabel(
            self.reg_delay_frame, 
            text="Delay Between Browsers", 
            font=self.font_label,
            text_color=self.colors["text_secondary"]
        )
        self.reg_delay_label.pack(side="left", padx=(0, 8))
        
        # Delay unit badge
        self.delay_unit_badge = ctk.CTkLabel(
            self.reg_delay_frame,
            text="seconds",
            font=self._font("Segoe UI", 10),
            text_color=self.colors["text_muted"]
        )
        self.delay_unit_badge.pack(side="left")
        
        self.reg_delay_entry = ctk.CTkEntry(
            self.reg_delay_frame, 
            placeholder_text="2", 
            width=80,
            height=36,
            font=self.font_label,
            fg_color=self.colors["bg_elevated"],
            border_color=self.colors["border_subtle"],
            corner_radius=10
        )
        self.reg_delay_entry.insert(0, "2")
        self.reg_delay_entry.pack(side="right")
        
        # Info tooltip for delay
        self.delay_info = ctk.CTkLabel(
            self.reg_delay_frame,
            text="ℹ️",
            font=self._font("Segoe UI", 14),
            text_color=self.colors["text_muted"]
        )
        self.delay_info.pack(side="right", padx=(0, 8))
        
        # ═══ ADVANCED OPTIONS CARD ═══
        adv_card = ctk.CTkFrame(
            tab, 
            fg_color=self.colors["bg_card"],
            corner_radius=16,
            border_width=1,
            border_color=self.colors["border_subtle"]
        )
        adv_card.pack(fill="x", padx=8, pady=(0, 8))
        
        # Advanced Header
        adv_header = ctk.CTkFrame(adv_card, fg_color="transparent")
        adv_header.pack(fill="x", padx=16, pady=(6, 4))
        
        ctk.CTkLabel(
            adv_header, 
            text="🎯  Advanced Options", 
            font=self._font("Segoe UI Semibold", 14, "bold"),
            text_color=self.colors["text_primary"]
        ).pack(side="left")
        
        ctk.CTkFrame(adv_card, height=1, fg_color=self.colors["border_subtle"]).pack(fill="x", padx=16)
        
        # ═══ ROW 1: Email Mode + Password (same row) ═══
        self.reg_email_pass_frame = ctk.CTkFrame(adv_card, fg_color="transparent")
        self.reg_email_pass_frame.pack(fill="x", padx=16, pady=(6, 4))
        
        ctk.CTkLabel(
            self.reg_email_pass_frame, 
            text="📧  Email:", 
            font=self.font_label,
            text_color=self.colors["text_secondary"]
        ).pack(side="left", padx=(0, 6))
        
        self.reg_email_mode_var = ctk.StringVar(value="TinyHost")
        self.reg_email_mode_menu = ctk.CTkOptionMenu(
            self.reg_email_pass_frame, 
            values=["TinyHost", "OAuth2"], 
            variable=self.reg_email_mode_var,
            command=self.on_email_mode_change,
            font=self.font_label,
            fg_color=self.colors["bg_elevated"],
            button_color=self.colors["accent_tertiary"],
            button_hover_color=self.colors["accent_primary"],
            dropdown_fg_color=self.colors["bg_elevated"],
            dropdown_hover_color=self.colors["bg_card_hover"],
            dropdown_text_color=self.colors["text_primary"],
            text_color=self.colors["text_primary"],
            width=100,
            height=32,
            corner_radius=8
        )
        self.reg_email_mode_menu.pack(side="left")
        
        # OAuth2 status label
        self.reg_oauth2_status = ctk.CTkLabel(
            self.reg_email_pass_frame,
            text="",
            font=self._font("Segoe UI", 11),
            text_color=self.colors["text_muted"]
        )
        self.reg_oauth2_status.pack(side="left", padx=(8, 0))
        
        # OAuth2 Refresh button
        self.reg_oauth2_refresh = ctk.CTkButton(
            self.reg_email_pass_frame,
            text="🔄",
            width=32,
            height=32,
            font=self._font("Segoe UI", 16),
            fg_color=self.colors["bg_elevated"],
            hover_color=self.colors["bg_card_hover"],
            corner_radius=6,
            command=self.refresh_oauth2_accounts
        )
        self.reg_oauth2_refresh.pack(side="left", padx=(4, 0))
        self.reg_oauth2_refresh.pack_forget()  # Hidden by default
        
        # Separator
        ctk.CTkLabel(
            self.reg_email_pass_frame,
            text="│",
            font=self._font("Segoe UI", 14),
            text_color=self.colors["border_subtle"]
        ).pack(side="left", padx=(12, 12))
        
        ctk.CTkLabel(
            self.reg_email_pass_frame, 
            text="🔑", 
            font=self.font_label,
            text_color=self.colors["text_secondary"]
        ).pack(side="left", padx=(0, 6))
        
        self.reg_password_var = ctk.StringVar(value=DEFAULT_PASSWORD)
        self.reg_password_entry = ctk.CTkEntry(
            self.reg_email_pass_frame,
            textvariable=self.reg_password_var,
            font=self.font_label,
            fg_color=self.colors["bg_elevated"],
            border_color=self.colors["border_subtle"],
            text_color=self.colors["text_primary"],
            width=140,
            height=32,
            corner_radius=8,
            show="•"
        )
        self.reg_password_entry.pack(side="left")
        
        self.password_visible = False
        self.reg_password_toggle = ctk.CTkButton(
            self.reg_email_pass_frame,
            text="👁",
            width=32,
            height=32,
            fg_color=self.colors["bg_elevated"],
            hover_color=self.colors["bg_card_hover"],
            corner_radius=8,
            command=self.toggle_password_visibility
        )
        self.reg_password_toggle.pack(side="left", padx=(4, 0))
        
        self.reg_password_save = ctk.CTkButton(
            self.reg_email_pass_frame,
            text="💾",
            width=32,
            height=32,
            fg_color=self.colors["accent_tertiary"],
            hover_color=self.colors["accent_primary"],
            corner_radius=8,
            font=self._font("Segoe UI", 14),
            command=self.save_password_to_file
        )
        self.reg_password_save.pack(side="left", padx=(4, 0))
        
        # ═══ ROW 2: Proxy (compact) ═══
        self.reg_proxy_frame = ctk.CTkFrame(adv_card, fg_color="transparent")
        self.reg_proxy_frame.pack(fill="x", padx=16, pady=(4, 4))
        
        self.reg_proxy_var = ctk.BooleanVar(value=PROXY_ENABLED)
        self.reg_proxy_switch = ctk.CTkSwitch(
            self.reg_proxy_frame,
            text="",
            variable=self.reg_proxy_var,
            font=self.font_label,
            text_color=self.colors["text_secondary"],
            fg_color=self.colors["bg_elevated"],
            progress_color=self.colors["accent_primary"],
            button_color=self.colors["text_primary"],
            button_hover_color=self.colors["accent_primary"],
            width=40,
            command=self.toggle_proxy_inputs
        )
        self.reg_proxy_switch.pack(side="left")
        
        ctk.CTkLabel(
            self.reg_proxy_frame,
            text="🌐 Proxy:",
            font=self.font_label,
            text_color=self.colors["text_secondary"]
        ).pack(side="left", padx=(6, 0))
        
        self.reg_proxy_string_var = ctk.StringVar(value=PROXY_STRING)
        self.reg_proxy_entry = ctk.CTkEntry(
            self.reg_proxy_frame,
            textvariable=self.reg_proxy_string_var,
            font=self._font("Consolas", 11),
            fg_color=self.colors["bg_elevated"],
            border_color=self.colors["border_subtle"],
            text_color=self.colors["text_primary"],
            placeholder_text="user:pass@host:port",
            width=280,
            height=32,
            corner_radius=8
        )
        self.reg_proxy_entry.pack(side="left", padx=(8, 0))
        
        self.reg_proxy_save = ctk.CTkButton(
            self.reg_proxy_frame,
            text="💾",
            width=36,
            height=32,
            fg_color=self.colors["accent_tertiary"],
            hover_color=self.colors["accent_primary"],
            corner_radius=8,
            font=self._font("Segoe UI", 14),
            command=self.save_proxy_to_file
        )
        self.reg_proxy_save.pack(side="left", padx=(6, 0))
        
        self.reg_proxy_info = ctk.CTkLabel(
            self.reg_proxy_frame,
            text="✓" if PROXY_ENABLED and PROXY_STRING else "",
            font=self._font("Segoe UI", 10),
            text_color=self.colors["accent_green"] if PROXY_ENABLED else self.colors["text_muted"]
        )
        self.reg_proxy_info.pack(side="left", padx=(4, 0))
        
        proxy_state = "normal" if PROXY_ENABLED else "disabled"
        self.reg_proxy_entry.configure(state=proxy_state)
        self.reg_proxy_save.configure(state=proxy_state)
        
        # ═══ ROW 3: Checkout + 2FA (same row) ═══
        self.reg_toggles_frame = ctk.CTkFrame(adv_card, fg_color="transparent")
        self.reg_toggles_frame.pack(fill="x", padx=16, pady=(4, 10))
        
        self.reg_checkout_var = ctk.BooleanVar(value=False)
        self.reg_checkout_switch = ctk.CTkSwitch(
            self.reg_toggles_frame, 
            text="  Checkout",
            variable=self.reg_checkout_var,
            font=self.font_label,
            text_color=self.colors["text_secondary"],
            fg_color=self.colors["bg_elevated"],
            progress_color=self.colors["accent_primary"],
            button_color=self.colors["text_primary"],
            button_hover_color=self.colors["accent_primary"],
            command=self.toggle_checkout_type
        )
        self.reg_checkout_switch.pack(side="left")
        
        # Checkout Type dropdown (inline)
        self.reg_checkout_type_frame = ctk.CTkFrame(self.reg_toggles_frame, fg_color="transparent")
        self.reg_checkout_type_frame.pack(side="left", padx=(8, 0))
        
        self.reg_checkout_type_var = ctk.StringVar(value="Plus")
        self.reg_checkout_type_dropdown = ctk.CTkOptionMenu(
            self.reg_checkout_type_frame,
            values=["Plus", "Business", "Both"],
            variable=self.reg_checkout_type_var,
            font=self.font_label,
            fg_color=self.colors["bg_elevated"],
            button_color=self.colors["accent_secondary"],
            button_hover_color=self.colors["accent_primary"],
            dropdown_fg_color=self.colors["bg_elevated"],
            dropdown_hover_color=self.colors["bg_card_hover"],
            dropdown_text_color=self.colors["text_primary"],
            text_color=self.colors["text_primary"],
            width=90,
            height=32,
            corner_radius=8
        )
        self.reg_checkout_type_dropdown.pack(side="left")
        self.reg_checkout_type_frame.pack_forget()  # Hidden initially
        
        # Separator
        ctk.CTkLabel(
            self.reg_toggles_frame,
            text="│",
            font=self._font("Segoe UI", 14),
            text_color=self.colors["border_subtle"]
        ).pack(side="left", padx=(16, 16))
        
        self.reg_2fa_var = ctk.BooleanVar(value=False)
        self.reg_2fa_switch = ctk.CTkSwitch(
            self.reg_toggles_frame,
            text="  2FA (TOTP)",
            variable=self.reg_2fa_var,
            font=self.font_label,
            text_color=self.colors["text_secondary"],
            fg_color=self.colors["bg_elevated"],
            progress_color=self.colors["accent_purple"],
            button_color=self.colors["text_primary"],
            button_hover_color=self.colors["accent_purple"],
        )
        self.reg_2fa_switch.pack(side="left")
        
        # ═══ ACTION BUTTONS ═══
        self.reg_btn_frame = ctk.CTkFrame(tab, fg_color="transparent")
        self.reg_btn_frame.pack(fill="x", padx=8, pady=(8, 4))
        
        # Start Button with glow effect
        self.reg_start_btn = GlowButton(
            self.reg_btn_frame, 
            text="▶  START REGISTRATION", 
            command=self.start_registration_thread, 
            fg_color=self.colors["accent_green"], 
            hover_color="#00dd88",
            text_color="#0a0a0a",
            height=52, 
            corner_radius=14,
            font=self._font("Segoe UI", 14, "bold"),
            glow_color=self.colors["accent_green"]
        )
        self.reg_start_btn.pack(side="left", fill="x", expand=True, padx=(0, 8))
        
        # Stop Button
        self.reg_stop_btn = ctk.CTkButton(
            self.reg_btn_frame, 
            text="⏹", 
            command=self.stop_process, 
            fg_color=self.colors["bg_card"], 
            border_width=2, 
            border_color=self.colors["error"], 
            text_color=self.colors["error"], 
            hover_color="#2a1515", 
            width=52,
            height=52, 
            corner_radius=14,
            font=self._font("Segoe UI", 18), 
            state="disabled"
        )
        self.reg_stop_btn.pack(side="right")

    def setup_checkout_tab(self):
        """Setup the Checkout Capture tab with account selection table"""
        tab = self.tabview.tab("💳 Checkout Capture")
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_rowconfigure(1, weight=1)
        
        # ═══ CONTROLS CARD ═══
        controls_card = ctk.CTkFrame(
            tab, 
            fg_color=self.colors["bg_card"],
            corner_radius=16,
            border_width=1,
            border_color=self.colors["border_subtle"]
        )
        controls_card.grid(row=0, column=0, padx=8, pady=(4, 8), sticky="ew")
        
        # Header row
        header_frame = ctk.CTkFrame(controls_card, fg_color="transparent")
        header_frame.pack(fill="x", padx=16, pady=(10, 8))
        
        ctk.CTkLabel(
            header_frame, 
            text="💳  Checkout Link Capture", 
            font=self._font("Segoe UI Semibold", 14, "bold"),
            text_color=self.colors["text_primary"]
        ).pack(side="left")
        
        # Refresh button
        self.checkout_refresh_btn = ctk.CTkButton(
            header_frame,
            text="🔄 Refresh",
            width=80,
            height=28,
            fg_color=self.colors["bg_elevated"],
            hover_color=self.colors["bg_card_hover"],
            corner_radius=8,
            font=self._font("Segoe UI", 11),
            command=self.load_checkout_accounts
        )
        self.checkout_refresh_btn.pack(side="right")
        
        ctk.CTkFrame(controls_card, height=1, fg_color=self.colors["border_subtle"]).pack(fill="x", padx=16)
        
        # Options row
        options_frame = ctk.CTkFrame(controls_card, fg_color="transparent")
        options_frame.pack(fill="x", padx=16, pady=10)
        
        # Checkout type
        ctk.CTkLabel(
            options_frame, 
            text="Capture Type:", 
            font=self.font_label,
            text_color=self.colors["text_secondary"]
        ).pack(side="left", padx=(0, 8))
        
        self.checkout_type_var = ctk.StringVar(value="Plus")
        self.checkout_type_menu = ctk.CTkOptionMenu(
            options_frame,
            values=["Plus", "Business", "Both"],
            variable=self.checkout_type_var,
            width=120,
            height=32,
            font=self.font_label,
            dropdown_font=self.font_label,
            fg_color=self.colors["accent_tertiary"],
            button_color=self.colors["accent_secondary"],
            button_hover_color=self.colors["accent_primary"],
            dropdown_fg_color=self.colors["bg_card"],
            dropdown_hover_color=self.colors["bg_card_hover"],
            corner_radius=10
        )
        self.checkout_type_menu.pack(side="left", padx=(0, 16))
        
        # Select All / Deselect All
        self.checkout_select_all_btn = ctk.CTkButton(
            options_frame,
            text="☑ Select All",
            width=90,
            height=28,
            fg_color=self.colors["bg_elevated"],
            hover_color=self.colors["accent_primary"],
            corner_radius=8,
            font=self._font("Segoe UI", 11),
            command=self.checkout_select_all
        )
        self.checkout_select_all_btn.pack(side="left", padx=(0, 4))
        
        self.checkout_deselect_all_btn = ctk.CTkButton(
            options_frame,
            text="☐ Deselect All",
            width=100,
            height=28,
            fg_color=self.colors["bg_elevated"],
            hover_color=self.colors["error"],
            corner_radius=8,
            font=self._font("Segoe UI", 11),
            command=self.checkout_deselect_all
        )
        self.checkout_deselect_all_btn.pack(side="left")
        
        # Account count label
        self.checkout_count_label = ctk.CTkLabel(
            options_frame,
            text="0 accounts | 0 selected",
            font=self._font("Segoe UI", 11),
            text_color=self.colors["text_muted"]
        )
        self.checkout_count_label.pack(side="right")
        
        # ═══ MULTITHREAD OPTIONS (always visible, but disabled when < 2 selected) ═══
        self.checkout_mt_frame = ctk.CTkFrame(controls_card, fg_color="transparent", height=40)
        self.checkout_mt_frame.pack(fill="x", padx=16, pady=(0, 8))
        self.checkout_mt_frame.pack_propagate(False)  # Fixed height to prevent resize
        
        # Multithread enable switch
        self.checkout_mt_var = ctk.BooleanVar(value=False)
        self.checkout_mt_switch = ctk.CTkSwitch(
            self.checkout_mt_frame,
            text="  Multithread Mode",
            variable=self.checkout_mt_var,
            font=self.font_label,
            text_color=self.colors["text_muted"],  # Start muted
            fg_color=self.colors["bg_elevated"],
            progress_color=self.colors["accent_orange"],
            button_color=self.colors["text_primary"],
            button_hover_color=self.colors["accent_orange"],
            command=self.toggle_checkout_multithread,
            state="disabled"  # Start disabled
        )
        self.checkout_mt_switch.pack(side="left", padx=(0, 16), pady=4)
        
        # Thread count label
        self.checkout_threads_label = ctk.CTkLabel(
            self.checkout_mt_frame,
            text="Threads:",
            font=self.font_label,
            text_color=self.colors["text_muted"]
        )
        self.checkout_threads_label.pack(side="left", padx=(0, 8), pady=4)
        
        self.checkout_threads_entry = ctk.CTkEntry(
            self.checkout_mt_frame,
            placeholder_text="2",
            width=50,
            height=32,
            font=self.font_label,
            fg_color=self.colors["bg_elevated"],
            border_color=self.colors["border_subtle"],
            corner_radius=8,
            state="disabled"
        )
        self.checkout_threads_entry.insert(0, "2")
        self.checkout_threads_entry.pack(side="left", pady=4)
        
        # Delay label and entry
        self.checkout_delay_label = ctk.CTkLabel(
            self.checkout_mt_frame,
            text="Delay:",
            font=self.font_label,
            text_color=self.colors["text_muted"]
        )
        self.checkout_delay_label.pack(side="left", padx=(16, 8), pady=4)
        
        self.checkout_delay_entry = ctk.CTkEntry(
            self.checkout_mt_frame,
            placeholder_text="3",
            width=50,
            height=32,
            font=self.font_label,
            fg_color=self.colors["bg_elevated"],
            border_color=self.colors["border_subtle"],
            corner_radius=8,
            state="disabled"
        )
        self.checkout_delay_entry.insert(0, "3")  # Default 3s delay
        self.checkout_delay_entry.pack(side="left", pady=4)
        
        self.checkout_delay_unit = ctk.CTkLabel(
            self.checkout_mt_frame,
            text="s",
            font=self.font_label,
            text_color=self.colors["text_muted"]
        )
        self.checkout_delay_unit.pack(side="left", padx=(2, 0), pady=4)
        
        # Max threads info
        self.checkout_mt_info = ctk.CTkLabel(
            self.checkout_mt_frame,
            text="(select 2+)",
            font=self._font("Segoe UI", 10),
            text_color=self.colors["text_muted"]
        )
        self.checkout_mt_info.pack(side="left", padx=(8, 0), pady=4)
        
        # Recommendation badge
        self.checkout_mt_badge = ctk.CTkLabel(
            self.checkout_mt_frame,
            text="💡 Rec: 2 threads, 3s delay",
            font=self._font("Segoe UI", 10),
            text_color=self.colors["accent_yellow"]
        )
        self.checkout_mt_badge.pack(side="right")
        
        # ═══ ACCOUNTS TABLE (ttk.Treeview — 1 widget replaces N*6 CTk widgets) ═══
        table_frame = ctk.CTkFrame(
            tab, 
            fg_color=self.colors["bg_card"],
            corner_radius=16,
            border_width=1,
            border_color=self.colors["border_subtle"]
        )
        table_frame.grid(row=1, column=0, padx=8, pady=(0, 8), sticky="nsew")
        table_frame.grid_columnconfigure(0, weight=1)
        table_frame.grid_rowconfigure(0, weight=1)
        
        # Configure dark theme for ttk widgets
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Checkout.Treeview",
            background="#111827",
            foreground="#f8fafc",
            fieldbackground="#111827",
            rowheight=30,
            font=("Segoe UI", 11)
        )
        style.configure("Checkout.Treeview.Heading",
            background="#1e293b",
            foreground="#94a3b8",
            font=("Segoe UI", 12, "bold")
        )
        style.map("Checkout.Treeview",
            background=[("selected", "#1e3a5f")],
            foreground=[("selected", "#00f0ff")]
        )

        
        # Create Treeview
        self.checkout_tree = ttk.Treeview(
            table_frame,
            columns=("email", "plus", "business", "sold"),
            show="tree headings",
            selectmode="extended",
            style="Checkout.Treeview",
            height=6
        )
        self.checkout_tree.grid(row=0, column=0, padx=8, pady=8, sticky="nsew")
        
        # Column config
        self.checkout_tree.column("#0", width=40, stretch=False)
        self.checkout_tree.column("email", width=250, stretch=True)
        self.checkout_tree.column("plus", width=50, stretch=False, anchor="center")
        self.checkout_tree.column("business", width=70, stretch=False, anchor="center")
        self.checkout_tree.column("sold", width=50, stretch=False, anchor="center")
        
        self.checkout_tree.heading("#0", text="✓")
        self.checkout_tree.heading("email", text="Email")
        self.checkout_tree.heading("plus", text="Plus")
        self.checkout_tree.heading("business", text="Business")
        self.checkout_tree.heading("sold", text="Sold")
        
        # Scrollbar
        tree_scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.checkout_tree.yview)
        tree_scrollbar.grid(row=0, column=1, sticky="ns", pady=8)
        self.checkout_tree.configure(yscrollcommand=tree_scrollbar.set)
        
        # Click to toggle selection
        self.checkout_tree.bind("<ButtonRelease-1>", self._toggle_checkout_row)
        
        # Tags for visual styling
        self.checkout_tree.tag_configure("sold", foreground="#ff6b4a")
        self.checkout_tree.tag_configure("selected", foreground="#00f0ff")
        self.checkout_tree.tag_configure("evenrow", background="#0f172a")
        self.checkout_tree.tag_configure("oddrow", background="#1a2332")
        
        # Store for account data (compatible with business logic)
        self.checkout_account_vars = []
        
        # ═══ ACTION BUTTONS ═══
        btn_frame = ctk.CTkFrame(tab, fg_color="transparent")
        btn_frame.grid(row=2, column=0, padx=8, pady=(0, 4), sticky="ew")
        
        self.checkout_start_btn = GlowButton(
            btn_frame, 
            text="💳  START CAPTURE", 
            command=self.start_checkout_capture_thread, 
            fg_color=self.colors["accent_orange"], 
            hover_color="#ff8855",
            text_color="#0a0a0a",
            height=48, 
            corner_radius=14,
            font=self._font("Segoe UI", 13, "bold"),
            glow_color=self.colors["accent_orange"]
        )
        self.checkout_start_btn.pack(side="left", fill="x", expand=True, padx=(0, 8))
        
        self.checkout_stop_btn = ctk.CTkButton(
            btn_frame, 
            text="⏹", 
            command=self.stop_process, 
            fg_color="transparent",
            hover_color=self.colors["error"],
            text_color=self.colors["text_secondary"],
            width=52, height=48,
            corner_radius=14,
            font=self._font("Segoe UI", 18),
            border_width=2,
            border_color=self.colors["border_subtle"],
            state="disabled"
        )
        self.checkout_stop_btn.pack(side="right")
        
        # Load accounts on init
        self.after(2000, self.load_checkout_accounts)  # Defer: render UI first
    
    def load_checkout_accounts(self):
        """Load accounts from Excel and display in treeview table"""
        # Clear existing data
        self.checkout_account_vars.clear()
        for item in self.checkout_tree.get_children():
            self.checkout_tree.delete(item)
        
        # Load accounts
        excel_file = "chatgpt.xlsx"
        if not os.path.exists(excel_file):
            self.checkout_count_label.configure(text="No chatgpt.xlsx found")
            return
        
        accounts = load_checkout_accounts(excel_file)
        
        if not accounts:
            self.checkout_count_label.configure(text="No accounts found in Excel")
            return
        
        # Insert rows into treeview (1 widget vs N*6 CTk widgets)
        for idx, account in enumerate(accounts):
            is_sold = account.get("is_sold", False)
            
            # Plus status
            if account["plus_url"] == "no Plus offer":
                plus_status = "⛔"
            elif account["plus_url"]:
                plus_status = "✅"
            else:
                plus_status = "❌"
            
            # Business/Sold status
            business_status = "✅" if account["business_url"] else "❌"
            sold_status = "✅" if is_sold else "❌"
            
            # Checkbox text (sold accounts show dash, others unchecked box)
            check_text = "—" if is_sold else "☐"
            
            email_text = account["email"][:35] + "..." if len(account["email"]) > 35 else account["email"]
            
            # Build tags: sold + alternating row color
            row_tag = "evenrow" if idx % 2 == 0 else "oddrow"
            tags = ("sold", row_tag) if is_sold else (row_tag,)
            
            item_id = self.checkout_tree.insert(
                "", "end",
                text=check_text,
                values=(email_text, plus_status, business_status, sold_status),
                tags=tags
            )
            
            # Create BooleanVar for compatibility with business logic
            var = ctk.BooleanVar(value=False)
            self.checkout_account_vars.append({
                "var": var,
                "account": account,
                "tree_id": item_id
            })
        
        self.update_checkout_selection_count()
    
    def update_checkout_selection_count(self):
        """Update the selection count label and enable/disable multithread options"""
        # Count only non-sold accounts
        total = sum(1 for item in self.checkout_account_vars if not item["account"].get("is_sold", False))
        selected = sum(1 for item in self.checkout_account_vars if item["var"].get() and not item["account"].get("is_sold", False))
        self.checkout_count_label.configure(text=f"{total} accounts | {selected} selected")
        
        # Enable multithread options when 2+ accounts selected (no pack/unpack to avoid resize animation)
        if selected >= 2:
            # Enable switch and update colors
            self.checkout_mt_switch.configure(
                state="normal",
                text_color=self.colors["text_secondary"]
            )
            # Update max threads info
            max_threads = min(5, selected)
            self.checkout_mt_info.configure(
                text=f"(max {max_threads})",
                text_color=self.colors["accent_primary"]
            )
            self.checkout_threads_label.configure(text_color=self.colors["text_secondary"])
            self.checkout_delay_label.configure(text_color=self.colors["text_secondary"])
        else:
            # Disable switch and mute colors
            self.checkout_mt_switch.configure(
                state="disabled",
                text_color=self.colors["text_muted"]
            )
            self.checkout_mt_var.set(False)
            self.checkout_threads_entry.configure(state="disabled")
            self.checkout_delay_entry.configure(state="disabled")
            self.checkout_mt_info.configure(
                text="(select 2+ accounts)",
                text_color=self.colors["text_muted"]
            )
            self.checkout_threads_label.configure(text_color=self.colors["text_muted"])
            self.checkout_delay_label.configure(text_color=self.colors["text_muted"])
    
    def toggle_checkout_multithread(self):
        """Toggle multithread mode for checkout capture"""
        if self.checkout_mt_var.get():
            self.checkout_threads_entry.configure(state="normal")
            self.checkout_delay_entry.configure(state="normal")
        else:
            self.checkout_threads_entry.configure(state="disabled")
            self.checkout_delay_entry.configure(state="disabled")
    

    def _toggle_checkout_row(self, event):
        """Toggle selection of clicked treeview row"""
        item = self.checkout_tree.identify_row(event.y)
        if not item:
            return
        
        # Find the account var for this item
        for acc in self.checkout_account_vars:
            if acc.get("tree_id") == item:
                if acc["account"].get("is_sold", False):
                    return  # Can't select sold accounts
                
                # Toggle
                new_val = not acc["var"].get()
                acc["var"].set(new_val)
                
                # Update checkbox text
                self.checkout_tree.item(item, text="✅" if new_val else "☐")
                
                # Update tag for visual
                tags = list(self.checkout_tree.item(item, "tags"))
                if new_val and "selected" not in tags:
                    tags.append("selected")
                elif not new_val and "selected" in tags:
                    tags.remove("selected")
                self.checkout_tree.item(item, tags=tuple(tags))
                break
        
        self.update_checkout_selection_count()
    
    def checkout_select_all(self):
        """Select all non-sold accounts"""
        for item in self.checkout_account_vars:
            if not item["account"].get("is_sold", False):
                item["var"].set(True)
                tree_id = item.get("tree_id")
                if tree_id:
                    self.checkout_tree.item(tree_id, text="✅")
        self.update_checkout_selection_count()
    
    def checkout_deselect_all(self):
        """Deselect all accounts"""
        for item in self.checkout_account_vars:
            item["var"].set(False)
            tree_id = item.get("tree_id")
            if tree_id:
                self.checkout_tree.item(tree_id, text="☐")
        self.update_checkout_selection_count()
    
    def start_checkout_capture_thread(self):
        """Start checkout capture in thread"""
        threading.Thread(target=self.run_checkout_capture).start()
    
    def run_checkout_capture(self):
        """Run checkout capture for selected accounts (sequential or multithread)"""
        # Get selected accounts (exclude sold accounts)
        selected = [item["account"] for item in self.checkout_account_vars 
                   if item["var"].get() and not item["account"].get("is_sold", False)]
        
        if not selected:
            print("⚠️ No accounts selected!")
            return
        
        self.lock_ui(True)
        self.stop_event.clear()
        TextRedirector._suppress_output = False  # Re-enable logging for new run
        self.update_stats(0, 0)
        
        checkout_type = self.checkout_type_var.get()
        excel_file = "chatgpt.xlsx"
        
        # Check if multithread mode is enabled
        use_multithread = self.checkout_mt_var.get() and len(selected) >= 2
        
        if use_multithread:
            try:
                threads = int(self.checkout_threads_entry.get())
                max_allowed = min(5, len(selected))
                threads = max(1, min(threads, max_allowed))
            except (ValueError, TypeError):
                threads = 2

            # Get delay between browsers
            try:
                thread_delay = int(self.checkout_delay_entry.get())
                thread_delay = max(1, min(thread_delay, 10))  # 1-10s
            except (ValueError, TypeError):
                thread_delay = 3  # Default 3s
            
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Starting Checkout Capture | Type: {checkout_type} | Accounts: {len(selected)} | Threads: {threads} | Delay: {thread_delay}s")
            self.update_status("RUNNING", self.colors["accent_orange"], f"Capturing {len(selected)} accounts with {threads} threads...")
            
            success_count = 0
            fail_count = 0
            count_lock = threading.Lock()
            
            def run_checkout_worker(thread_id, account, account_idx, start_delay):
                nonlocal success_count, fail_count
                
                # Use pre-calculated start delay
                if start_delay > 0:
                    safe_print(thread_id, f"Waiting {start_delay}s before starting...", Colors.INFO, "⏳ ")
                    # Sleep in 0.5s intervals to check stop_event
                    for _ in range(int(start_delay * 2)):
                        if self.stop_event.is_set():
                            return False, account["email"]
                        time.sleep(0.5)
                
                if self.stop_event.is_set():
                    return False, account["email"]
                
                worker = CheckoutCaptureWorker(
                    thread_id=thread_id,
                    email=account["email"],
                    access_token=account["access_token"],
                    excel_file=excel_file,
                    row_index=account["row_index"],
                    checkout_type=checkout_type
                )
                worker.stop_event = self.stop_event

                self._register_worker(worker)
                try:
                    result = worker.run()
                finally:
                    self._unregister_worker(worker)
                
                with count_lock:
                    if result:
                        nonlocal success_count
                        success_count += 1
                        print(f"✅ Captured: {account['email']}")
                    else:
                        nonlocal fail_count
                        fail_count += 1
                        print(f"❌ Failed: {account['email']}")
                    self.update_stats(success_count, fail_count)
                
                return result, account["email"]
            
            # Run with ThreadPoolExecutor
            executor = concurrent.futures.ThreadPoolExecutor(max_workers=threads)
            self._register_executor(executor)
            try:
                futures = []
                for idx, account in enumerate(selected):
                    if self.stop_event.is_set():
                        break
                    
                    # Slot index is (idx % threads) for staggered delay in each wave
                    slot_idx = idx % threads
                    wave_idx = idx // threads
                    
                    if wave_idx == 0:
                        # First wave: stagger by slot
                        start_delay = slot_idx * thread_delay
                    else:
                        # Subsequent waves: Base delay (5s) + stagger
                        start_delay = 5.0 + (slot_idx * thread_delay)
                    
                    thread_id = (idx % threads) + 1
                    future = executor.submit(run_checkout_worker, thread_id, account, idx, start_delay)
                    futures.append(future)
                
                # Wait for all futures
                for future in concurrent.futures.as_completed(futures):
                    if self.stop_event.is_set():
                        executor.shutdown(wait=False, cancel_futures=True)
                        break
                    try:
                        future.result()
                    except Exception as e:
                        print(f"Error in worker: {e}")
            finally:
                try:
                    executor.shutdown(wait=False, cancel_futures=True)
                except Exception:
                    pass
                self._unregister_executor(executor)
        else:
            # Sequential mode
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Starting Checkout Capture | Type: {checkout_type} | Accounts: {len(selected)}")
            self.update_status("RUNNING", self.colors["accent_orange"], f"Capturing {len(selected)} accounts...")
            
            success_count = 0
            fail_count = 0
            
            for idx, account in enumerate(selected):
                if self.stop_event.is_set():
                    break
                
                self.update_status("RUNNING", self.colors["accent_orange"], f"Account {idx+1}/{len(selected)}: {account['email']}")
                
                worker = CheckoutCaptureWorker(
                    thread_id=1,
                    email=account["email"],
                    access_token=account["access_token"],
                    excel_file=excel_file,
                    row_index=account["row_index"],
                    checkout_type=checkout_type
                )
                worker.stop_event = self.stop_event

                self._register_worker(worker)
                try:
                    result = worker.run()
                finally:
                    self._unregister_worker(worker)
                
                if result:
                    success_count += 1
                    print(f"✅ Captured: {account['email']}")
                else:
                    fail_count += 1
                    print(f"❌ Failed: {account['email']}")
                
                self.update_stats(success_count, fail_count)
        
        # Schedule UI updates on main thread to avoid Tkinter threading issues
        final_success = success_count
        final_fail = fail_count
        
        def finish_capture():
            # Stage 1: Lightweight UI updates (immediate, keeps GUI responsive)
            self.lock_ui(False)
            if self.stop_event.is_set():
                self.update_status("STOPPED", self.colors["warning"], f"Stopped. Success: {final_success} | Failed: {final_fail}")
            else:
                # Use neutral color for COMPLETED (like IDLE)
                self.update_status("COMPLETED", None, f"Success: {final_success} | Failed: {final_fail}")
            
            print(f"💳 Captured {final_success} checkout links!")

            # Stage 2: Schedule heavy table rebuild with delay so GUI stays responsive
            self.after(300, self.load_checkout_accounts)
        
        self.after(100, finish_capture)

    # --- UI ANIMATIONS ---
    def _blink_cursor(self):
        """Animate terminal cursor blinking"""
        try:
            if not hasattr(self, '_cursor_visible'):
                self._cursor_visible = True
            
            self._cursor_visible = not self._cursor_visible
            if self._cursor_visible:
                self.cursor_indicator.configure(text_color=self.colors["accent_primary"])
            else:
                self.cursor_indicator.configure(text_color=self.colors["glass_bg"])
            
            self.after(530, self._blink_cursor)
        except Exception:
            pass
    
    # --- UI LOGIC ---
    def toggle_reg_inputs(self, choice):
        if choice == "Sequential":
            self.reg_count_label.configure(text="Account Count")
            # Hide threads and delay inputs for Sequential mode
            self.reg_threads_frame.pack_forget()
            self.reg_delay_frame.pack_forget()
        else:
            self.reg_count_label.configure(text="Total")
            # Show threads inline with count, delay on separate row
            self.reg_threads_frame.pack(side="right", padx=(20, 0))
            self.reg_delay_frame.pack(fill="x", padx=16, pady=(0, 10), after=self.reg_count_frame)

    def toggle_checkout_type(self):
        """Show/hide checkout type dropdown based on switch state"""
        if self.reg_checkout_var.get():
            self.reg_checkout_type_frame.pack(side="left", padx=(20, 0))
        else:
            self.reg_checkout_type_frame.pack_forget()
    
    def on_email_mode_change(self, mode):
        """Handle email mode change between TinyHost and OAuth2"""
        global oauth2_accounts
        
        if mode == "OAuth2":
            # Show refresh button and load accounts
            self.reg_oauth2_refresh.pack(side="left", padx=(8, 0))
            self._load_oauth2_and_update_status()
        else:
            # TinyHost mode - clear status and hide refresh button
            self.reg_oauth2_status.configure(text="")
            self.reg_oauth2_refresh.pack_forget()
            oauth2_accounts = []

    def refresh_oauth2_accounts(self):
        """Refresh OAuth2 accounts from oauth2.xlsx"""
        self._load_oauth2_and_update_status()

    def _load_oauth2_and_update_status(self):
        """Load OAuth2 accounts and update status label"""
        global oauth2_accounts
        excel_file = "oauth2.xlsx"
        if os.path.exists(excel_file):
            oauth2_accounts = load_oauth2_accounts_from_excel(excel_file)
            count = len(oauth2_accounts)
            if count > 0:
                self.reg_oauth2_status.configure(
                    text=f"✅ Loaded {count} OAuth2 accounts",
                    text_color=self.colors["accent_green"]
                )
            else:
                self.reg_oauth2_status.configure(
                    text="⚠️ No OAuth2 accounts found",
                    text_color=self.colors["warning"]
                )
        else:
            self.reg_oauth2_status.configure(
                text="❌ oauth2.xlsx not found",
                text_color=self.colors["error"]
            )
            oauth2_accounts = []
    
    def toggle_password_visibility(self):
        """Toggle password visibility"""
        self.password_visible = not self.password_visible
        if self.password_visible:
            self.reg_password_entry.configure(show="")
            self.reg_password_toggle.configure(text="🙈")
        else:
            self.reg_password_entry.configure(show="•")
            self.reg_password_toggle.configure(text="👁")
    
    def save_password_to_file(self):
        """Save password to the source code file"""
        new_password = self.reg_password_var.get().strip()
        if not new_password:
            print("❌ Password cannot be empty!")
            return
        
        try:
            # Get the current file path
            current_file = os.path.abspath(__file__)
            
            # Read the file content
            with open(current_file, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # Match: DEFAULT_PASSWORD = "anything"
            pattern = r'DEFAULT_PASSWORD\s*=\s*"[^"]*"'
            replacement = f'DEFAULT_PASSWORD = "{new_password}"'
            
            # Check if pattern exists
            if not re.search(pattern, content):
                print("❌ Could not find DEFAULT_PASSWORD in code!")
                return
            
            new_content = re.sub(pattern, replacement, content, count=1)
            
            # Verify the replacement was made
            if new_content == content:
                print("❌ Password replacement failed!")
                return
            
            # Write back to file
            with open(current_file, 'w', encoding='utf-8') as f:
                f.write(new_content)
            
            # Update global variable
            global DEFAULT_PASSWORD
            DEFAULT_PASSWORD = new_password
            
            print("✅ Password saved successfully")
            
        except Exception as e:
            print(f"❌ Failed to save: {str(e)}")
    
    def toggle_proxy_inputs(self):
        """Enable/disable proxy inputs based on switch state"""
        global PROXY_ENABLED
        enabled = self.reg_proxy_var.get()
        PROXY_ENABLED = enabled
        state = "normal" if enabled else "disabled"
        self.reg_proxy_entry.configure(state=state)
        self.reg_proxy_save.configure(state=state)
        
        if enabled:
            self.reg_proxy_info.configure(text="ON", text_color=self.colors["accent_green"])
        else:
            self.reg_proxy_info.configure(text="", text_color=self.colors["text_muted"])
    
    def save_proxy_to_file(self):
        """Save proxy configuration to JSON file"""
        global PROXY_ENABLED, PROXY_STRING, PROXY_FORMAT
        
        proxy_string = self.reg_proxy_string_var.get().strip()
        enabled = self.reg_proxy_var.get()
        
        if not proxy_string:
            print("❌ Proxy string cannot be empty!")
            return
        
        # Auto-detect format and validate
        detected = detect_proxy_format(proxy_string)
        if not detected:
            print("❌ Cannot detect proxy format! Use user:pass@host:port or host:port:user:pass")
            return
        
        proxy_info, urls = parse_proxy(proxy_string, detected)
        if not proxy_info:
            print("❌ Invalid proxy! Check your input.")
            return
        
        # Save to config file
        if save_proxy_config(enabled, proxy_string, detected):
            PROXY_ENABLED = enabled
            PROXY_STRING = proxy_string
            PROXY_FORMAT = detected
            
            print(f"✅ Proxy saved: {proxy_info['host']}:{proxy_info['port']}")
            self.reg_proxy_info.configure(text="✓", text_color=self.colors["accent_green"])
            
            # Reset info text after 2 seconds
            self.after(2000, lambda: self.reg_proxy_info.configure(
                text="ON" if enabled else "",
                text_color=self.colors["accent_green"] if enabled else self.colors["text_muted"]
            ))
        else:
            print("❌ Failed to save proxy config!")
            
    def update_status(self, state="IDLE", color=None, details=""):
        # Thread-safety: schedule on main thread if called from background thread
        if threading.current_thread() is not threading.main_thread():
            self.after(0, lambda: self.update_status(state, color, details))
            return
        # Map state to icon (keep consistent neutral color for all states)
        state_icons = {
            "IDLE": "●",
            "RUNNING": "◉",
            "LOADING": "◎",
            "COMPLETED": "✓",
            "STOPPED": "◼",
            "ERROR": "✗"
        }
        
        icon = state_icons.get(state.upper(), "●")
        self.status_indicator.configure(text=f"{icon} {state.upper()}")
        
        # Use consistent neutral color for all states (like IDLE)
        neutral_color = self.colors["bg_elevated"]
        neutral_border = self.colors["border_subtle"]
        self.status_indicator.configure(fg_color=neutral_color, border_color=neutral_border)
        
        # Update status dot - keep neutral for all states
        st = state.upper()
        if st == "RUNNING":
            self.status_dot.base_color = self.colors["text_muted"]
            self.status_dot.start_pulse()
        else:
            self.status_dot.stop_pulse()
            self.status_dot.configure(fg_color=self.colors["text_muted"])
            
        if details:
            self.info_box.configure(state="normal")
            self.info_box.delete("0.0", "end")
            self.info_box.insert("0.0", f"→ {details}")
            self.info_box.configure(state="disabled")
            
    def update_stats(self, success, failed):
        """Update stats cards with animated counters"""
        # Thread-safety: schedule on main thread if called from background thread
        if threading.current_thread() is not threading.main_thread():
            self.after(0, lambda: self.update_stats(success, failed))
            return
        # Direct counter updates (no animation)
        self.success_count_label.configure(text=str(success))
        self.fail_count_label.configure(text=str(failed))
        
        # Update progress percentage
        total = success + failed
        if total > 0:
            percent = int((success / total) * 100) if total > 0 else 0
            self.progress_percent.configure(text=f"{percent}% success rate")

    def lock_ui(self, is_running):
        # Thread-safety: schedule on main thread if called from background thread
        if threading.current_thread() is not threading.main_thread():
            self.after(0, lambda: self.lock_ui(is_running))
            return
        self.running = is_running
        state = "disabled" if is_running else "normal"
        stop_state = "normal" if is_running else "disabled"
        
        # Update buttons with animated states
        if is_running:
            self.reg_start_btn.configure(
                state=state, 
                text="⏳  PROCESSING...",
                fg_color=self.colors["bg_elevated"]
            )
            self.checkout_start_btn.configure(
                state=state, 
                text="⏳  PROCESSING...",
                fg_color=self.colors["bg_elevated"]
            )
        else:
            # Only re-enable logging if not forcefully stopped
            # (workers may still be dying and would flood output)
            if not self.stop_event.is_set():
                TextRedirector._suppress_output = False
            self.reg_start_btn.configure(
                state=state, 
                text="▶  START REGISTRATION",
                fg_color=self.colors["accent_green"]
            )
            self.checkout_start_btn.configure(
                state=state, 
                text="💳  START CAPTURE",
                fg_color=self.colors["accent_orange"]
            )
        
        # Stop buttons
        self.reg_stop_btn.configure(state=stop_state)
        self.checkout_stop_btn.configure(state=stop_state)
        
        
        # Progress Bar (custom 60fps animation)
        if is_running:
            self.progress_bar.set(0.5)
            self.update_status("RUNNING", self.colors["info"], "Initializing automation...")
        else:
            self.progress_bar.set(0)
            
        # Lock inputs
        self.reg_mode_menu.configure(state=state)

        self.reg_count_entry.configure(state=state)
        self.reg_threads_entry.configure(state=state)
        self.reg_delay_entry.configure(state=state)
        self.reg_checkout_switch.configure(state=state)
        self.reg_checkout_type_dropdown.configure(state=state)
        self.reg_2fa_switch.configure(state=state)
        self.checkout_type_menu.configure(state=state)
        self.checkout_refresh_btn.configure(state=state)
        self.checkout_select_all_btn.configure(state=state)
        self.checkout_deselect_all_btn.configure(state=state)
        self.checkout_mt_switch.configure(state=state)
        if not is_running and self.checkout_mt_var.get():
            self.checkout_threads_entry.configure(state="normal")
            self.checkout_delay_entry.configure(state="normal")
        else:
            self.checkout_threads_entry.configure(state="disabled" if is_running else ("normal" if self.checkout_mt_var.get() else "disabled"))
            self.checkout_delay_entry.configure(state="disabled" if is_running else ("normal" if self.checkout_mt_var.get() else "disabled"))
        self.reg_password_entry.configure(state=state)
        self.reg_password_toggle.configure(state=state)
        self.reg_password_save.configure(state=state)
        
        # Lock proxy inputs
        self.reg_proxy_switch.configure(state=state)
        if is_running:
            self.reg_proxy_entry.configure(state="disabled")
            self.reg_proxy_save.configure(state="disabled")
        else:
            proxy_state = "normal" if self.reg_proxy_var.get() else "disabled"
            self.reg_proxy_entry.configure(state=proxy_state)
            self.reg_proxy_save.configure(state=proxy_state)
        
        

    def stop_process(self):
        TextRedirector._suppress_output = True  # Suppress ALL log output immediately
        self.stop_event.set()
        self.update_status("STOPPING", self.colors["warning"], "Force stopping now...")
        # Run heavy cleanup in background thread to avoid freezing GUI
        threading.Thread(target=self._background_force_stop, daemon=True).start()
        print("⏹ Force stop sent (closing browsers now)")

    def _background_force_stop(self):
        """Run heavy cleanup off the main thread to keep GUI responsive"""
        self._force_stop_all_runtime()
        self._kill_chromium_processes()
    
    @staticmethod
    def _kill_chromium_processes():
        """Kill any orphaned browser processes spawned by patchright"""
        _kill_browser_processes()
    
    def on_closing(self):
        """Clean shutdown: kill all browsers, release files, then destroy window"""
        try:
            self.stop_event.set()
            self._force_stop_all_runtime()
            self._kill_chromium_processes()
        except Exception:
            pass
        finally:
            self.destroy()
            os._exit(0)  # Force exit to kill any remaining threads

    def _register_worker(self, worker):
        if not worker:
            return
        with self._active_runtime_lock:
            self._active_workers.add(worker)

    def _unregister_worker(self, worker):
        if not worker:
            return
        with self._active_runtime_lock:
            self._active_workers.discard(worker)

    def _register_executor(self, executor):
        if not executor:
            return
        with self._active_runtime_lock:
            self._active_executors.add(executor)

    def _unregister_executor(self, executor):
        if not executor:
            return
        with self._active_runtime_lock:
            self._active_executors.discard(executor)

    def _force_stop_worker(self, worker):
        """Force stop a worker by killing browser PID directly (fast, non-blocking)"""
        if not worker:
            return
        try:
            if hasattr(worker, 'stop_event'):
                worker.stop_event = self.stop_event
        except Exception:
            pass

        # Kill browser via Playwright server PID (kills server + all child browsers)
        pw_server_pid = getattr(worker, '_pw_server_pid', None)
        if pw_server_pid:
            try:
                subprocess.Popen(
                    ['taskkill', '/F', '/PID', str(pw_server_pid), '/T'],
                    stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                    creationflags=subprocess.CREATE_NO_WINDOW
                )
            except Exception:
                pass

        # Nullify references (don't call .close() — it hangs)
        try:
            worker.page = None
            worker.context = None
            worker.browser = None
            worker.playwright = None
            worker._pw_server_pid = None
        except Exception:
            pass

        # Skip playwright.stop() during force stop — it hangs when browser is dead.
        # taskkill /IM chromium.exe handles all cleanup.
        try:
            worker.playwright = None
        except Exception:
            pass

        # Stop proxy bridge if any
        try:
            bridge = getattr(worker, 'proxy_bridge', None)
            if bridge:
                bridge.stop()
                worker.proxy_bridge = None
        except Exception:
            pass

    def _force_stop_all_runtime(self):
        with self._active_runtime_lock:
            executors = list(self._active_executors)
            workers = list(self._active_workers)

        for executor in executors:
            try:
                executor.shutdown(wait=False, cancel_futures=True)
            except Exception:
                pass

        # Kill all workers in parallel (don't let one hanging worker delay others)
        kill_threads = []
        for worker in workers:
            t = threading.Thread(target=self._force_stop_worker, args=(worker,), daemon=True)
            t.start()
            kill_threads.append(t)
        # Wait max 3 seconds for all kills to complete
        for t in kill_threads:
            t.join(timeout=3)

    # --- LOG UTILS ---
    def clear_logs(self):
        self.log_textbox.configure(state="normal")
        self.log_textbox.delete("0.0", "end")
        self.log_textbox.configure(state="disabled")
        print("🗑️ Logs cleared!")
        
    def copy_logs(self):
        try:
            text = self.log_textbox.get("0.0", "end")
            self.clipboard_clear()
            self.clipboard_append(text)
            print("📋 Copied to clipboard!")
        except Exception:
            pass
            
    def export_logs(self):
        try:
            filename = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text Files", "*.txt")])
            if filename:
                with open(filename, "w", encoding="utf-8") as f:
                    f.write(self.log_textbox.get("0.0", "end"))
                print("💾 Exported successfully!")
        except Exception as e:
            print(f"❌ Export failed!")
            messagebox.showerror("Error", f"Could not export logs: {e}")

    # --- WORKERS ---
    def start_registration_thread(self):
        threading.Thread(target=self.run_registration).start()

    def run_registration(self):
        global GET_CHECKOUT_LINK, GET_CHECKOUT_TYPE, ENABLE_2FA, oauth2_accounts, PROXY_ENABLED, PROXY_STRING, PROXY_FORMAT

        # Reset stop state and re-enable logging for new run
        self.stop_event.clear()
        TextRedirector._suppress_output = False

        self.lock_ui(True)
        self.update_stats(0, 0)
        GET_CHECKOUT_LINK = self.reg_checkout_var.get()
        GET_CHECKOUT_TYPE = self.reg_checkout_type_var.get()
        ENABLE_2FA = self.reg_2fa_var.get()
        
        # Apply proxy settings from GUI
        PROXY_ENABLED = self.reg_proxy_var.get()
        PROXY_STRING = self.reg_proxy_string_var.get().strip()
        if PROXY_ENABLED and PROXY_STRING:
            detected_fmt = detect_proxy_format(PROXY_STRING)
            if detected_fmt:
                PROXY_FORMAT = detected_fmt
            proxy_info, _ = parse_proxy(PROXY_STRING)
            if proxy_info:
                print(f"🌐 Proxy enabled: {proxy_info['host']}:{proxy_info['port']}")
            else:
                print(f"⚠️ Invalid proxy string, running without proxy")
                PROXY_ENABLED = False
        
        email_mode = self.reg_email_mode_var.get()  # "TinyHost" or "OAuth2"
        mode = self.reg_mode_var.get()
        try:
            count = int(self.reg_count_entry.get())
        except (ValueError, TypeError):
            count = 1
        
        # OAuth2 mode validation
        if email_mode == "OAuth2":
            # Reload oauth2 accounts to get fresh list
            excel_file = "oauth2.xlsx"
            if os.path.exists(excel_file):
                oauth2_accounts = load_oauth2_accounts_from_excel(excel_file)
            else:
                oauth2_accounts = []
            
            available_count = len(oauth2_accounts)
            if available_count == 0:
                print(f"❌ No OAuth2 accounts available in oauth2.xlsx")
                self.lock_ui(False)
                return
            
            if count > available_count:
                print(f"⚠️ Requested {count} accounts but only {available_count} OAuth2 accounts available. Using {available_count}.")
                count = available_count
            
            # Reset used status
            reset_oauth2_accounts()
            print(f"📧 Using OAuth2 mode with {available_count} available accounts")
            
        # Get thread count (only for multithread mode)
        try:
            threads = int(self.reg_threads_entry.get())
            if threads < 1:
                threads = 2
            if threads > count:
                threads = count  # Can't have more threads than accounts
        except (ValueError, TypeError):
            threads = 2

        # Get thread delay (only for multithread mode)
        try:
            thread_delay = float(self.reg_delay_entry.get())
            if thread_delay < 0:
                thread_delay = 2
        except (ValueError, TypeError):
            thread_delay = 2
            
        if mode == "Multithread":
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Starting Registration | Mode: {mode} | Email: {email_mode} | Total: {count} | Threads: {threads} | Delay: {thread_delay}s")
        else:
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Starting Registration | Mode: {mode} | Email: {email_mode} | Count: {count}")
        
        success_count = 0
        failed_count = 0
        
        if mode == "Sequential":
            for i in range(count):
                if self.stop_event.is_set():
                    break
                
                self.update_status("RUNNING", self.colors["info"], f"Processing Account {i+1}/{count}")
                print(f"--- Account {i+1}/{count} ---")
                
                try:
                    # Get oauth2 account if in OAuth2 mode
                    oauth2_account = None
                    if email_mode == "OAuth2":
                        oauth2_account = get_next_oauth2_account()
                        if not oauth2_account:
                            print(f"❌ No more OAuth2 accounts available")
                            break
                    
                    worker = ChatGPTAutoRegisterWorker(
                        thread_id=1, 
                        email_mode=email_mode, 
                        oauth2_account=oauth2_account
                    )
                    worker.stop_event = self.stop_event
                    self._register_worker(worker)
                    try:
                        success, result = worker.run()
                    finally:
                        self._unregister_worker(worker)
                    if success:
                        success_count += 1
                        print("✅ Success!")
                        # Mark OAuth2 account as registered
                        if email_mode == "OAuth2" and oauth2_account:
                            mark_oauth2_registered(oauth2_account["row_num"])
                    else:
                        failed_count += 1
                        print("❌ Failed!")
                except Exception as e:
                    failed_count += 1
                    print(f"❌ Error: {e}")

                
                self.update_stats(success_count, failed_count)
                
                if i < count - 1 and not self.stop_event.is_set():
                    print("Waiting 3 seconds...")
                    time.sleep(3)
        else:
            # Multithread with fixed number of threads
            self.update_status("RUNNING", self.colors["info"], f"Processing {count} accounts with {threads} threads (delay: {thread_delay}s)...")
            
            processed = 0
            
            # Create a wrapper function that calculates delay based on slot position
            def run_worker_with_slot_delay(account_idx, slot_idx, stop_event, start_delay, email_mode_inner, num_threads_inner):
                """Worker with computed start delay"""
                # Delay is pre-calculated
                if start_delay > 0:
                    safe_print(account_idx, f"Waiting {start_delay}s before starting...", Colors.INFO, "⏳ ")
                    for _ in range(int(start_delay * 2)):
                        if stop_event and stop_event.is_set():
                            return (False, None, None)
                        time.sleep(0.5)
                
                # Get oauth2 account if in OAuth2 mode
                oauth2_account = None
                oauth2_row_num = None
                if email_mode_inner == "OAuth2":
                    oauth2_account = get_next_oauth2_account()
                    if not oauth2_account:
                        safe_print(account_idx, "No more OAuth2 accounts available", Colors.ERROR, "❌ ")
                        return (False, None, None)
                    oauth2_row_num = oauth2_account.get("row_num")
                
                # Use slot_idx+1 as thread_id for proper window positioning
                worker = ChatGPTAutoRegisterWorker(
                    slot_idx + 1,  # Use slot index for window position (1-based)
                    num_threads=num_threads_inner,
                    email_mode=email_mode_inner,
                    oauth2_account=oauth2_account
                )
                worker.stop_event = stop_event
                self._register_worker(worker)
                try:
                    success, result = worker.run()
                finally:
                    self._unregister_worker(worker)
                
                # Mark OAuth2 account as registered if success
                if success and email_mode_inner == "OAuth2" and oauth2_row_num:
                    mark_oauth2_registered(oauth2_row_num)
                
                return (success, result, oauth2_row_num)


            
            executor = ThreadPoolExecutor(max_workers=threads)
            self._register_executor(executor)
            try:
                # Submit tasks
                futures = {}
                for i in range(count):
                    if self.stop_event.is_set():
                        break
                    
                    # Slot index is (i % threads) for staggered delay in each wave
                    slot_idx = i % threads
                    wave_idx = i // threads
                    
                    if wave_idx == 0:
                        # First wave: stagger by slot
                        start_delay = slot_idx * thread_delay
                    else:
                        # Subsequent waves: Base delay (5s) + stagger
                        # This ensures threads wait 5s + stagger before starting next account
                        start_delay = 5.0 + (slot_idx * thread_delay)
                    
                    future = executor.submit(run_worker_with_slot_delay, i+1, slot_idx, self.stop_event, start_delay, email_mode, threads)
                    futures[future] = i+1
                
                for future in as_completed(futures):
                    if self.stop_event.is_set():
                        executor.shutdown(wait=False, cancel_futures=True)
                        break
                    try:
                        result_tuple = future.result()
                        # Handle both 2-tuple (TinyHost) and 3-tuple (OAuth2) returns
                        if len(result_tuple) == 3:
                            success, result, _ = result_tuple
                        else:
                            success, result = result_tuple
                        processed += 1
                        if success:
                            success_count += 1
                        else:
                            failed_count += 1
                        self.update_stats(success_count, failed_count)
                        self.update_status("RUNNING", self.colors["info"], f"Processed {processed}/{count} accounts")
                    except Exception as e:
                        failed_count += 1
                        processed += 1
                        self.update_stats(success_count, failed_count)
                        print(f"Error: {e}")
                    
                    self.update_stats(success_count, failed_count)
            finally:
                try:
                    executor.shutdown(wait=False, cancel_futures=True)
                except Exception:
                    pass
                self._unregister_executor(executor)

        final_msg = "COMPLETED" if not self.stop_event.is_set() else "STOPPED"
        color = self.colors["success"] if not self.stop_event.is_set() else self.colors["warning"]
        self.update_status(final_msg, color, f"✨ Success: {success_count} | Failed: {failed_count}")
        print(f"\n{'🎉' if not self.stop_event.is_set() else '🛑'} {final_msg}! Success: {success_count} | Failed: {failed_count}")
        
        # Show completion toast (schedule on main thread)
        if not self.stop_event.is_set():
            self.after(0, lambda: print(f"✅ Completed! {success_count} accounts"))
        else:
            self.after(0, lambda: print(f"⏹ Stopped. {success_count} completed"))
        
        self.lock_ui(False)
        # Restore status with neutral color (like IDLE) - schedule on main thread
        stopped = self.stop_event.is_set()
        self.after(0, lambda: self.status_indicator.configure(
            text=f"{'✓' if not stopped else '◼'} {final_msg}",
            fg_color=self.colors["bg_elevated"],
            border_color=self.colors["border_subtle"]
        ))


def _kill_browser_processes():
    """Kill any orphaned browser processes spawned by patchright"""
    for proc_name in ['chrome.exe', 'chromium.exe']:
        try:
            subprocess.Popen(
                ['taskkill', '/F', '/IM', proc_name, '/T'],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                creationflags=subprocess.CREATE_NO_WINDOW
            )
        except Exception:
            pass

atexit.register(_kill_browser_processes)

if __name__ == "__main__":
    app = App()
    app.protocol("WM_DELETE_WINDOW", app.on_closing)
    app.mainloop()
