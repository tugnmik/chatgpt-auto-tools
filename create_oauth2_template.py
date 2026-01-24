"""
Script tạo file oauth2.xlsx template
File này dùng để lưu tài khoản OAuth2 riêng biệt với chatgpt.xlsx

Cấu trúc:
- Cột A: email|password|refresh_token|token_id
- Cột B: Status (registered = đã đăng ký, trống = chưa đăng ký)
"""

from openpyxl import Workbook
import os

def create_oauth2_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "OAuth2 Accounts"
    
    # Headers
    ws['A1'] = "email|password|refresh_token|token_id"
    ws['B1'] = "Status"
    
    # Set column widths
    ws.column_dimensions['A'].width = 80
    ws.column_dimensions['B'].width = 15
    
    # Example row (commented format)
    # ws['A2'] = "example@outlook.com|password123|refresh_token_here|token_id_here"
    # ws['B2'] = ""  # Empty = not registered yet
    
    output_file = "oauth2.xlsx"
    
    if os.path.exists(output_file):
        print(f"⚠️ File {output_file} already exists! Rename or delete it first.")
        return False
    
    wb.save(output_file)
    print(f"✅ Created {output_file}")
    print("📝 Add your OAuth2 accounts in column A (format: email|password|refresh_token|token_id)")
    print("📌 Status column B will be auto-filled with 'registered' after successful registration")
    return True

if __name__ == "__main__":
    create_oauth2_template()
